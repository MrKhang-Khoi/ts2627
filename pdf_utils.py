import os
import zipfile
import io
import shutil
from datetime import datetime
from pypdf import PdfWriter, PdfReader
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

_BASE = os.environ.get('DATA_DIR', os.path.dirname(__file__))
UPLOAD_FOLDER = os.path.join(_BASE, 'uploads')
BACKUP_FOLDER = os.path.join(_BASE, 'backups')

ALLOWED_EXTENSIONS = {'pdf', 'jpg', 'jpeg', 'png'}


# ===== HELPER: lấy bytes từ local path HOẶC Google Drive =====

def _get_bytes(file_path):
    """
    Lấy bytes của file từ local path hoặc Google Drive.
    Returns (bytes, error).
    """
    if not file_path:
        return None, 'file_path trống'
    try:
        import drive_utils
        if drive_utils.is_drive(file_path):
            return drive_utils.download_bytes(file_path)
    except ImportError:
        pass
    # Local file
    if os.path.exists(file_path):
        with open(file_path, 'rb') as f:
            return f.read(), None
    return None, f'File không tồn tại: {file_path}'


def _save_or_upload(pdf_bytes, dest_path, student, doc_type):
    """
    Lưu PDF bytes vào local path HOẶC upload lên Drive.
    Returns (file_path, error).
    """
    try:
        import drive_utils
        if drive_utils.DRIVE_MODE:
            return drive_utils.upload(pdf_bytes, student, doc_type)
    except ImportError:
        pass
    # Local
    try:
        os.makedirs(os.path.dirname(dest_path), exist_ok=True)
        with open(dest_path, 'wb') as f:
            f.write(pdf_bytes)
        return dest_path, None
    except Exception as e:
        return None, f'Lỗi lưu file: {str(e)}'


# ===== MERGE MULTIPLE FILES (trả về bytes) =====

def merge_to_bytes(file_objects, max_mb=5):
    """
    Nhận danh sách Werkzeug FileStorage, merge thành PDF bytes.
    Dùng cho upload-multi (CCCD 2 mặt...).
    Returns (pdf_bytes, page_count, error).
    """
    if not file_objects:
        return None, 0, 'Chưa chọn file nào.'
    writer = PdfWriter()
    for f in file_objects:
        f.seek(0, 2); size = f.tell(); f.seek(0)
        if size == 0:
            return None, 0, f'File "{getattr(f,"filename","?")}" rỗng.'
        if size > max_mb * 1024 * 1024:
            return None, 0, f'File "{getattr(f,"filename","?")}" quá lớn (tối đa {max_mb}MB).'
        ext = (getattr(f, 'filename', '') or '').rsplit('.', 1)[-1].lower()
        if ext not in ALLOWED_EXTENSIONS:
            return None, 0, f'File "{getattr(f,"filename","?")}" không hợp lệ.'
        try:
            if ext == 'pdf':
                reader = PdfReader(f)
                if len(reader.pages) == 0:
                    return None, 0, f'PDF "{getattr(f,"filename","?")}" không có trang.'
                for page in reader.pages:
                    writer.add_page(page)
            else:
                from PIL import Image
                img = Image.open(f)
                if img.mode != 'RGB':
                    img = img.convert('RGB')
                tmp = io.BytesIO()
                img.save(tmp, 'PDF', resolution=150)
                tmp.seek(0)
                for page in PdfReader(tmp).pages:
                    writer.add_page(page)
        except Exception as e:
            return None, 0, f'Lỗi đọc file "{getattr(f,"filename","?")}": {str(e)}'
    if len(writer.pages) == 0:
        return None, 0, 'Không có trang nào được tạo.'
    buf = io.BytesIO()
    writer.write(buf)
    return buf.getvalue(), len(writer.pages), None


def merge_multiple_pdfs(file_objects, dest_path, backup_dir=None, doc_type='FILE', max_mb=20):
    """
    Backward-compatible wrapper — dùng merge_to_bytes rồi lưu vào disk.
    Không dùng trong Drive mode (các route sẽ gọi merge_to_bytes trực tiếp).
    """
    pdf_bytes, _, err = merge_to_bytes(file_objects, max_mb)
    if err:
        return None, err
    # Backup file cũ
    if backup_dir and os.path.exists(dest_path):
        os.makedirs(backup_dir, exist_ok=True)
        ts = datetime.now().strftime('%Y-%m-%d_%H%M%S')
        shutil.copy2(dest_path, os.path.join(backup_dir, f'backup_{doc_type}_{ts}.pdf'))
    try:
        os.makedirs(os.path.dirname(dest_path), exist_ok=True)
        with open(dest_path, 'wb') as out:
            out.write(pdf_bytes)
        return dest_path, None
    except Exception as e:
        return None, f'Lỗi lưu file: {str(e)}'


# ===== APPEND TO EXISTING PDF =====

def append_bytes(existing_bytes_or_none, new_file_objects, max_mb=5):
    """
    Thêm các file mới vào cuối existing PDF bytes.
    Returns (pdf_bytes, page_count, error).
    """
    writer = PdfWriter()
    # Đọc file hiện tại (nếu có)
    if existing_bytes_or_none:
        try:
            reader = PdfReader(io.BytesIO(existing_bytes_or_none))
            for page in reader.pages:
                writer.add_page(page)
        except Exception as e:
            return None, 0, f'Lỗi đọc file hiện tại: {str(e)}'
    # Thêm file mới
    for f in new_file_objects:
        f.seek(0, 2); size = f.tell(); f.seek(0)
        if size == 0:
            return None, 0, f'File "{getattr(f,"filename","?")}" rỗng.'
        if size > max_mb * 1024 * 1024:
            return None, 0, f'File quá lớn (tối đa {max_mb}MB).'
        ext = (getattr(f, 'filename', '') or '').rsplit('.', 1)[-1].lower()
        if ext not in ALLOWED_EXTENSIONS:
            return None, 0, f'File "{getattr(f,"filename","?")}" không hợp lệ.'
        try:
            if ext == 'pdf':
                for page in PdfReader(f).pages:
                    writer.add_page(page)
            else:
                from PIL import Image
                img = Image.open(f)
                if img.mode != 'RGB':
                    img = img.convert('RGB')
                tmp = io.BytesIO()
                img.save(tmp, 'PDF', resolution=150)
                tmp.seek(0)
                for page in PdfReader(tmp).pages:
                    writer.add_page(page)
        except Exception as e:
            return None, 0, f'Lỗi đọc file: {str(e)}'
    if len(writer.pages) == 0:
        return None, 0, 'Không có trang nào được tạo.'
    buf = io.BytesIO()
    writer.write(buf)
    return buf.getvalue(), len(writer.pages), None


def append_to_existing_pdf(existing_path, new_file_objects, dest_path, backup_dir=None, max_mb=20):
    """
    Backward-compatible — đọc existing từ disk, gọi append_bytes, lưu vào disk.
    Không dùng trong Drive mode.
    """
    existing_bytes = None
    if existing_path and os.path.exists(existing_path):
        with open(existing_path, 'rb') as f:
            existing_bytes = f.read()
    pdf_bytes, page_count, err = append_bytes(existing_bytes, new_file_objects, max_mb)
    if err:
        return None, 0, err
    if backup_dir and os.path.exists(dest_path):
        os.makedirs(backup_dir, exist_ok=True)
        ts = datetime.now().strftime('%Y-%m-%d_%H%M%S')
        shutil.copy2(dest_path, os.path.join(backup_dir, f'backup_HOCBA_{ts}.pdf'))
    try:
        os.makedirs(os.path.dirname(dest_path), exist_ok=True)
        with open(dest_path, 'wb') as out:
            out.write(pdf_bytes)
        return dest_path, page_count, None
    except Exception as e:
        return None, 0, f'Lỗi lưu file: {str(e)}'


# ===== MERGE TRANSCRIPTS (học bạ hoàn chỉnh) =====

def merge_transcripts(student_folder, ma_hoso, doc_map):
    """
    Nối học bạ 6-8 và học bạ 9 thành học bạ hoàn chỉnh.
    Hỗ trợ cả local path và Drive file.
    """
    path_6_8 = doc_map.get('HOCBA_6_8', {}).get('file_path')
    path_9   = doc_map.get('HOCBA_9',   {}).get('file_path')

    errors = []
    bytes_6_8, err = _get_bytes(path_6_8) if path_6_8 else (None, 'Thiếu file HOCBA_6_8')
    if err: errors.append(err)
    bytes_9, err   = _get_bytes(path_9)   if path_9   else (None, 'Thiếu file HOCBA_9')
    if err: errors.append(err)

    if errors:
        return None, '; '.join(errors)

    try:
        writer = PdfWriter()
        for b in [bytes_6_8, bytes_9]:
            reader = PdfReader(io.BytesIO(b))
            for page in reader.pages:
                writer.add_page(page)
        buf = io.BytesIO()
        writer.write(buf)
        merged_bytes = buf.getvalue()
    except Exception as e:
        return None, f'Lỗi khi nối PDF: {str(e)}'

    # Lưu vào Drive hoặc local
    try:
        import drive_utils
        if drive_utils.DRIVE_MODE:
            # Tạo student dict giả từ doc_map context
            # Cần lop và ho_ten_khong_dau — lấy từ folder path
            parts = student_folder.replace('\\', '/').split('/')
            lop_guess = parts[-2] if len(parts) >= 2 else 'UNKNOWN'
            # Upload HOCBA.pdf
            fake_student = {'lop': lop_guess, 'ma_hoso': ma_hoso, 'ho_ten_khong_dau': parts[-1].replace(ma_hoso + '_', '')}
            return drive_utils.upload(merged_bytes, fake_student, 'HOCBA')
    except ImportError:
        pass

    # Local
    dest = os.path.join(student_folder, 'HOCBA.pdf')
    if os.path.exists(dest):
        ts = datetime.now().strftime('%Y-%m-%d_%H%M%S')
        bk_dir = os.path.join(BACKUP_FOLDER, ma_hoso)
        os.makedirs(bk_dir, exist_ok=True)
        shutil.copy2(dest, os.path.join(bk_dir, f'backup_HOCBA_{ts}.pdf'))
    with open(dest, 'wb') as f:
        f.write(merged_bytes)
    return dest, None


# ===== EXPORT EXCEL =====

def export_excel(students_data, class_name=None):
    """Xuất danh sách hồ sơ ra Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = class_name or 'Toàn khối'

    header_fill = PatternFill("solid", fgColor="1565C0")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ['STT', 'Lớp', 'Họ tên', 'Ngày sinh', 'Giấy khai sinh',
               'CCCD', 'Học bạ 6-8', 'Học bạ 9', 'Học bạ HT', 'CN Tốt nghiệp', 'Trạng thái', 'Ghi chú', 'Cập nhật lần cuối']
    ws.append(headers)
    for col, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    status_labels = {
        'CHUA_NOP': 'Chưa nộp', 'DA_NOP_CHO_KIEM_TRA': 'Chờ KT',
        'DAT': 'Đạt', 'FILE_MO': 'File mờ', 'SAI_FILE': 'Sai file',
        'THIEU_TRANG': 'Thiếu trang', 'CAN_NOP_LAI': 'Cần nộp lại', 'DA_KHOA': 'Đã khóa',
    }
    overall_labels = {
        'CHUA_NOP': 'Chưa nộp', 'TAM_DU_GIAI_DOAN_1': 'Tạm đủ GĐ1',
        'CHUA_DU': 'Chưa đủ', 'CAN_SUA': 'Cần sửa', 'DU_HO_SO_CHINH_THUC': 'Đủ hồ sơ',
    }
    doc_keys = ['GIAYKHAISINH', 'CCCD', 'HOCBA_6_8', 'HOCBA_9', 'HOCBA', 'CNTN_THCS']

    for i, s in enumerate(students_data, 1):
        docs = s.get('docs', {})
        row = [
            s.get('stt', i), s.get('lop'), s.get('ho_ten'), s.get('ngay_sinh'),
        ]
        for dk in doc_keys:
            st = docs.get(dk, {}).get('status', 'CHUA_NOP')
            row.append(status_labels.get(st, st))
        row.append(overall_labels.get(s.get('status_overall', ''), s.get('status_overall', '')))
        row.append(s.get('note', ''))
        row.append(s.get('updated_at', '')[:10] if s.get('updated_at') else '')
        ws.append(row)
        for cell in ws[i + 1]:
            cell.border = border
            cell.alignment = center

    col_widths = [6, 8, 22, 12, 14, 12, 12, 10, 12, 16, 16, 20, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 30

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ===== ZIP CREATION (hỗ trợ cả Drive và local) =====

def _get_hocba_path(doc_map):
    """Lấy đường dẫn học bạ: ưu tiên HOCBA hoàn chỉnh, fallback HOCBA_6_8."""
    fp = doc_map.get('HOCBA', {}).get('file_path')
    if fp:
        return fp, 'HOCBA'
    fp = doc_map.get('HOCBA_6_8', {}).get('file_path')
    if fp:
        return fp, 'HOCBA_6_8'
    return None, None


def _add_to_zip(zf, file_path, arcname):
    """Thêm file vào ZIP — hỗ trợ cả local và Drive."""
    if not file_path:
        return
    b, err = _get_bytes(file_path)
    if err or not b:
        return  # Bỏ qua file lỗi, không crash ZIP
    zf.writestr(arcname, b)


def create_student_zip(student, doc_map):
    """Tạo ZIP hồ sơ của một học sinh với đúng thứ tự nộp."""
    buf = io.BytesIO()
    hocba_path, _ = _get_hocba_path(doc_map)
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        ordered = [
            (doc_map.get('GIAYKHAISINH', {}).get('file_path'), '01_GiayKhaiSinh.pdf'),
            (doc_map.get('CNTN_THCS',    {}).get('file_path'), '02_ChungNhanTotNghiep.pdf'),
            (hocba_path,                                        '03_HocBa.pdf'),
            (doc_map.get('CCCD',         {}).get('file_path'), '04_CCCD.pdf'),
            (doc_map.get('ANH_THE',      {}).get('file_path'), '05_AnhThe.pdf'),
            (doc_map.get('UU_TIEN',      {}).get('file_path'), '06_UuTien.pdf'),
        ]
        for fp, fname in ordered:
            _add_to_zip(zf, fp, fname)
    buf.seek(0)
    return buf


def create_class_zip(class_name, students_data):
    """Tạo ZIP hồ sơ cả lớp với đúng thứ tự."""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for s in students_data:
            folder_name = f"{s['stt']:02d}_{s['ho_ten_khong_dau']}" if s.get('stt') else s['ho_ten_khong_dau']
            docs = s.get('docs', {})
            hocba_path, _ = _get_hocba_path(docs)
            ordered = [
                (docs.get('GIAYKHAISINH', {}).get('file_path'), '01_GiayKhaiSinh.pdf'),
                (docs.get('CNTN_THCS',    {}).get('file_path'), '02_ChungNhanTotNghiep.pdf'),
                (hocba_path,                                     '03_HocBa.pdf'),
                (docs.get('CCCD',         {}).get('file_path'), '04_CCCD.pdf'),
                (docs.get('ANH_THE',      {}).get('file_path'), '05_AnhThe.pdf'),
                (docs.get('UU_TIEN',      {}).get('file_path'), '06_UuTien.pdf'),
            ]
            for fp, fname in ordered:
                _add_to_zip(zf, fp, f"{class_name}/{folder_name}/{fname}")
    buf.seek(0)
    return buf


def create_all_zip(all_students):
    """Tạo ZIP toàn khối với đúng thứ tự."""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for s in all_students:
            folder_name = f"{s['stt']:02d}_{s['ho_ten_khong_dau']}" if s.get('stt') else s['ho_ten_khong_dau']
            docs = s.get('docs', {})
            hocba_path, _ = _get_hocba_path(docs)
            ordered = [
                (docs.get('GIAYKHAISINH', {}).get('file_path'), '01_GiayKhaiSinh.pdf'),
                (docs.get('CNTN_THCS',    {}).get('file_path'), '02_ChungNhanTotNghiep.pdf'),
                (hocba_path,                                     '03_HocBa.pdf'),
                (docs.get('CCCD',         {}).get('file_path'), '04_CCCD.pdf'),
                (docs.get('ANH_THE',      {}).get('file_path'), '05_AnhThe.pdf'),
                (docs.get('UU_TIEN',      {}).get('file_path'), '06_UuTien.pdf'),
            ]
            for fp, fname in ordered:
                _add_to_zip(zf, fp, f"{s['lop']}/{folder_name}/{fname}")
    buf.seek(0)
    return buf
