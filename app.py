from flask import Flask, render_template, request, redirect, url_for, session, jsonify, send_file, flash
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
import os, io, shutil, openpyxl, threading, time as _time_mod
from datetime import datetime
from database import get_db, init_db, add_log, get_setting, set_setting, update_overall_status
from file_utils import (to_ascii, get_student_folder, save_uploaded_file, delete_file_to_backup,
                        DOC_TYPES, DOC_LABELS, STATUS_LABELS, OVERALL_LABELS, UPLOAD_FOLDER,
                        DISPLAY_ORDER, OPTIONAL_DOCS, MULTI_FILE_DOCS)
from pdf_utils import (merge_transcripts, export_excel, create_student_zip,
                       create_class_zip, create_all_zip, append_to_existing_pdf)

app = Flask(__name__)
# Ã„ÂÃ¡Â»Âc secret key tÃ¡Â»Â« biÃ¡ÂºÂ¿n mÃƒÂ´i trÃ†Â°Ã¡Â»Âng (bÃ¡ÂºÂ¯t buÃ¡Â»â„¢c trÃƒÂªn Render)
app.secret_key = os.environ.get('SECRET_KEY', 'hoso_lop10_secret_key_2026_local')
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB Ã¢â‚¬â€ file lÃƒÂªn Drive, khÃƒÂ´ng lo quota disk

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get('role') != 'admin':
            flash('BÃ¡ÂºÂ¡n khÃƒÂ´ng cÃƒÂ³ quyÃ¡Â»Ân truy cÃ¡ÂºÂ­p trang nÃƒÂ y.', 'error')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated

def get_student_docs(student_id):
    conn = get_db()
    rows = conn.execute("SELECT * FROM documents WHERE student_id=?", (student_id,)).fetchall()
    conn.close()
    doc_map = {}
    for r in rows:
        doc_map[r['doc_type']] = dict(r)
    return doc_map

def get_student_with_docs(student_id):
    conn = get_db()
    s = conn.execute("SELECT * FROM students WHERE id=?", (student_id,)).fetchone()
    conn.close()
    if not s:
        return None, {}
    return dict(s), get_student_docs(student_id)

def enrich_students(students):
    """ThÃƒÂªm docs vÃƒÂ  doc_count cho mÃ¡Â»â€”i hÃ¡Â»Âc sinh."""
    result = []
    for s in students:
        sd = dict(s)
        docs = get_student_docs(s['id'])
        sd['docs'] = docs
        # Ã„ÂÃ¡ÂºÂ¿m sÃ¡Â»â€˜ tÃƒÂ i liÃ¡Â»â€¡u Ã„â€˜ÃƒÂ£ nÃ¡Â»â„¢p (cÃƒÂ³ file, khÃƒÂ´ng tÃƒÂ­nh CHUA_NOP)
        sd['doc_count'] = sum(1 for d in docs.values() if d.get('file_path'))
        sd['doc_total'] = len(DISPLAY_ORDER)  # tÃ¡Â»â€¢ng sÃ¡Â»â€˜ mÃ¡Â»Â¥c
        result.append(sd)
    return result

@app.route('/')
def index():
    conn = get_db()
    classes = [r['lop'] for r in conn.execute("SELECT DISTINCT lop FROM students ORDER BY lop").fetchall()]
    conn.close()
    # Kiem tra moi truong: PythonAnywhere khong co Playwright => an nut TSDC
    import socket
    hostname = socket.gethostname().lower()
    is_pythonanywhere = 'pythonanywhere' in hostname or os.path.exists('/home/ts102627')
    show_tsdc = not is_pythonanywhere
    return render_template('index.html', classes=classes, show_tsdc=show_tsdc)

@app.route('/huong-dan')
def huong_dan():
    """Trang hÃ†Â°Ã¡Â»â€ºng dÃ¡ÂºÂ«n nÃ¡Â»â„¢p hÃ¡Â»â€œ sÃ†Â¡ dÃƒÂ nh cho hÃ¡Â»Âc sinh"""
    return send_file('static/huong_dan_hoso.html')

@app.route('/class/<class_name>')
def student_list(class_name):
    conn = get_db()
    students = conn.execute("SELECT * FROM students WHERE lop=? ORDER BY CAST(stt AS INTEGER)", (class_name,)).fetchall()
    conn.close()
    students = enrich_students(students)
    return render_template('student_list.html', students=students, class_name=class_name,
                           DOC_LABELS=DOC_LABELS, STATUS_LABELS=STATUS_LABELS, OVERALL_LABELS=OVERALL_LABELS)

@app.route('/student/<int:student_id>')
def student_profile(student_id):
    student, doc_map = get_student_with_docs(student_id)
    if not student:
        flash('KhÃƒÂ´ng tÃƒÂ¬m thÃ¡ÂºÂ¥y hÃ¡Â»Âc sinh.', 'error')
        return redirect(url_for('index'))
    # BÃ¡ÂºÂ£o mÃ¡ÂºÂ­t: hÃ¡Â»Âc sinh phÃ¡ÂºÂ£i xÃƒÂ¡c minh mÃƒÂ£ CCCD trÃ†Â°Ã¡Â»â€ºc khi xem hÃ¡Â»â€œ sÃ†Â¡
    if session.get('role') not in ('teacher', 'admin'):
        verified = session.get('verified_students', [])
        if student_id not in verified:
            # Redirect vÃ¡Â»Â trang lÃ¡Â»â€ºp, front-end sÃ¡ÂºÂ½ mÃ¡Â»Å¸ modal xÃƒÂ¡c minh
            return redirect(url_for('student_list', class_name=student['lop']) + f'?verify={student_id}')
    phase = get_setting('phase', '1')
    max_mb = int(get_setting('max_file_size_mb', '20'))
    active_docs = list(DISPLAY_ORDER)
    return render_template('student_profile.html', student=student, doc_map=doc_map,
                           DOC_TYPES=DOC_TYPES, DOC_LABELS=DOC_LABELS, STATUS_LABELS=STATUS_LABELS,
                           OVERALL_LABELS=OVERALL_LABELS, active_docs=active_docs, phase=phase,
                           max_mb=max_mb, DISPLAY_ORDER=DISPLAY_ORDER, OPTIONAL_DOCS=OPTIONAL_DOCS,
                           MULTI_FILE_DOCS=MULTI_FILE_DOCS)

@app.route('/api/download-hoso/<int:student_id>')
def api_download_hoso(student_id):
    """TÃ¡ÂºÂ£i toÃƒÂ n bÃ¡Â»â„¢ hÃ¡Â»â€œ sÃ†Â¡ cÃ¡Â»Â§a mÃ¡Â»â„¢t hÃ¡Â»Âc sinh dÃ¡ÂºÂ¡ng ZIP"""
    student, doc_map = get_student_with_docs(student_id)
    if not student:
        flash('KhÃƒÂ´ng tÃƒÂ¬m thÃ¡ÂºÂ¥y hÃ¡Â»Âc sinh.', 'error')
        return redirect(url_for('index'))
    buf = create_student_zip(student, doc_map)
    filename = f"HoSo_{student['ho_ten_khong_dau']}_{student['lop']}.zip"
    return send_file(buf, mimetype='application/zip',
                     as_attachment=True, download_name=filename)
@app.route('/api/verify-student-code', methods=['POST'])
def api_verify_student_code():
    """HÃ¡Â»Âc sinh nhÃ¡ÂºÂ­p CCCD Ã„â€˜Ã¡Â»Æ’ xÃƒÂ¡c minh danh tÃƒÂ­nh trÃ†Â°Ã¡Â»â€ºc khi xem hÃ¡Â»â€œ sÃ†Â¡."""
    data = request.get_json(force=True) or {}
    student_id = data.get('student_id')
    code = (data.get('code') or '').strip().replace(' ', '').replace('-', '')
    if not student_id or not code:
        return jsonify({'error': 'Vui lÃƒÂ²ng nhÃ¡ÂºÂ­p mÃƒÂ£ bÃ¡ÂºÂ£o mÃ¡ÂºÂ­t.'}), 400
    conn = get_db()
    row = conn.execute("SELECT id, ma_hoso, ho_ten, lop FROM students WHERE id=?", (student_id,)).fetchone()
    conn.close()
    if not row:
        return jsonify({'error': 'KhÃƒÂ´ng tÃƒÂ¬m thÃ¡ÂºÂ¥y hÃ¡Â»Âc sinh.'}), 404
    # Normalize: chuÃ¡ÂºÂ©n hoÃƒÂ¡ vÃ¡Â»Â 12 chÃ¡Â»Â¯ sÃ¡Â»â€˜ Ã„â€˜Ã¡Â»Æ’ xÃ¡Â»Â­ lÃƒÂ½ trÃ†Â°Ã¡Â»Âng hÃ¡Â»Â£p
    # Excel lÃ†Â°u CCCD lÃƒÂ  sÃ¡Â»â€˜ (mÃ¡ÂºÂ¥t sÃ¡Â»â€˜ 0 Ã„â€˜Ã¡ÂºÂ§u, vÃƒÂ­ dÃ¡Â»Â¥: 064311001234 Ã¢â€ â€™ 64311001234)
    code_norm    = code.zfill(12)
    db_norm      = str(row['ma_hoso']).strip().replace(' ', '').zfill(12)
    if code_norm == db_norm:
        verified = session.get('verified_students', [])
        sid = int(student_id)
        if sid not in verified:
            verified.append(sid)
        session['verified_students'] = verified
        session.modified = True
        add_log(None, 'STUDENT_VERIFY', student_id, None, f"HS xÃƒÂ¡c minh: {row['ho_ten']}")
        return jsonify({'success': True,
                        'redirect': url_for('student_profile', student_id=student_id)})
    else:
        add_log(None, 'VERIFY_FAIL', student_id, None,
                f"Sai mÃƒÂ£: {row['ho_ten']} | nhÃ¡ÂºÂ­p={code_norm} | db={db_norm}")
        return jsonify({'error': 'MÃƒÂ£ bÃ¡ÂºÂ£o mÃ¡ÂºÂ­t khÃƒÂ´ng Ã„â€˜ÃƒÂºng. KiÃ¡Â»Æ’m tra lÃ¡ÂºÂ¡i sÃ¡Â»â€˜ CCCD cÃ¡Â»Â§a em.'}), 401


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        conn = get_db()
        user = conn.execute("SELECT * FROM users WHERE username=?", (username,)).fetchone()
        conn.close()
        if user and check_password_hash(user['password_hash'], password):
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['role'] = user['role']
            session['full_name'] = user['full_name'] or username
            session['assigned_class'] = user['assigned_class']
            if user['role'] == 'admin':
                return redirect(url_for('admin_dashboard'))
            return redirect(url_for('teacher_dashboard'))
        flash('TÃƒÂªn Ã„â€˜Ã„Æ’ng nhÃ¡ÂºÂ­p hoÃ¡ÂºÂ·c mÃ¡ÂºÂ­t khÃ¡ÂºÂ©u khÃƒÂ´ng Ã„â€˜ÃƒÂºng.', 'error')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/teacher')
@login_required
def teacher_dashboard():
    assigned = session.get('assigned_class')
    conn = get_db()
    if session.get('role') == 'admin':
        classes = [r['lop'] for r in conn.execute("SELECT DISTINCT lop FROM students ORDER BY lop").fetchall()]
        selected = request.args.get('class', classes[0] if classes else None)
    else:
        classes = [assigned] if assigned else []
        selected = assigned
    students = []
    if selected:
        rows = conn.execute("SELECT * FROM students WHERE lop=? ORDER BY CAST(stt AS INTEGER)", (selected,)).fetchall()
        students = enrich_students(rows)
    conn.close()
    stats = compute_stats(students)
    return render_template('teacher_dashboard.html', students=students, classes=classes,
                           selected_class=selected, stats=stats, DOC_LABELS=DOC_LABELS,
                           STATUS_LABELS=STATUS_LABELS, OVERALL_LABELS=OVERALL_LABELS, DOC_TYPES=DOC_TYPES)

@app.route('/admin')
@login_required
@admin_required
def admin_dashboard():
    conn = get_db()
    classes = [r['lop'] for r in conn.execute("SELECT DISTINCT lop FROM students ORDER BY lop").fetchall()]
    selected = request.args.get('class', 'ALL')
    filter_status = request.args.get('filter', '')
    query = "SELECT * FROM students"
    params = []
    if selected and selected != 'ALL':
        query += " WHERE lop=?"
        params.append(selected)
    query += " ORDER BY lop, CAST(stt AS INTEGER)"
    students = enrich_students(conn.execute(query, params).fetchall())
    if filter_status:
        students = [s for s in students if s['status_overall'] == filter_status]
    teachers = conn.execute("SELECT * FROM users WHERE role='teacher' ORDER BY full_name").fetchall()
    phase = get_setting('phase', '1')
    max_mb = get_setting('max_file_size_mb', '20')
    conn.close()
    stats = compute_stats(students)
    return render_template('admin_dashboard.html', students=students, classes=classes,
                           selected_class=selected, filter_status=filter_status, stats=stats,
                           teachers=teachers, phase=phase, max_mb=max_mb,
                           DOC_LABELS=DOC_LABELS, STATUS_LABELS=STATUS_LABELS,
                           OVERALL_LABELS=OVERALL_LABELS, DOC_TYPES=DOC_TYPES)

def compute_stats(students):
    total = len(students)
    stats = {'total': total, 'DU_HO_SO_CHINH_THUC': 0, 'CHUA_DU': 0, 'CAN_SUA': 0,
             'CHUA_NOP': 0, 'TAM_DU_GIAI_DOAN_1': 0}
    for s in students:
        key = s.get('status_overall', 'CHUA_NOP')
        if key in stats:
            stats[key] += 1
        else:
            stats['CHUA_DU'] += 1
    return stats

# ===== API ROUTES =====
@app.route('/api/classes')
def api_classes():
    conn = get_db()
    classes = [r['lop'] for r in conn.execute("SELECT DISTINCT lop FROM students ORDER BY lop").fetchall()]
    conn.close()
    return jsonify(classes)

@app.route('/api/students')
def api_students():
    lop = request.args.get('class', '')
    conn = get_db()
    rows = conn.execute("SELECT * FROM students WHERE lop=? ORDER BY stt", (lop,)).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/student/<int:sid>')
def api_student(sid):
    student, doc_map = get_student_with_docs(sid)
    if not student:
        return jsonify({'error': 'KhÃƒÂ´ng tÃƒÂ¬m thÃ¡ÂºÂ¥y'}), 404
    return jsonify({'student': student, 'docs': doc_map})

@app.route('/api/upload', methods=['POST'])
def api_upload():
    student_id = request.form.get('student_id', type=int)
    doc_type = request.form.get('doc_type', '').upper()
    file = request.files.get('file')
    if not student_id or doc_type not in DOC_TYPES:
        return jsonify({'error': 'DÃ¡Â»Â¯ liÃ¡Â»â€¡u khÃƒÂ´ng hÃ¡Â»Â£p lÃ¡Â»â€¡.'}), 400
    student, doc_map = get_student_with_docs(student_id)
    if not student:
        return jsonify({'error': 'KhÃƒÂ´ng tÃƒÂ¬m thÃ¡ÂºÂ¥y hÃ¡Â»Âc sinh.'}), 404
    # KiÃ¡Â»Æ’m tra khÃƒÂ³a
    existing = doc_map.get(doc_type, {})
    if existing.get('locked') and existing.get('status') == 'DAT':
        return jsonify({'error': 'File Ã„â€˜ÃƒÂ£ Ã„â€˜Ã†Â°Ã¡Â»Â£c Ã„â€˜ÃƒÂ¡nh dÃ¡ÂºÂ¥u Ã„ÂÃ¡ÂºÂ¡t. LiÃƒÂªn hÃ¡Â»â€¡ giÃƒÂ¡o viÃƒÂªn Ã„â€˜Ã¡Â»Æ’ mÃ¡Â»Å¸ khÃƒÂ³a.'}), 403
    max_mb = int(get_setting('max_file_size_mb', '20'))
    file_path, err = save_uploaded_file(file, student, doc_type, max_mb)
    if err:
        return jsonify({'error': err}), 400
    conn = get_db()
    now = datetime.now().isoformat()
    conn.execute("DELETE FROM documents WHERE student_id=? AND doc_type=?", (student_id, doc_type))
    conn.execute("""INSERT INTO documents (student_id,doc_type,file_name,file_path,status,uploaded_at)
                    VALUES (?,?,?,?,?,?)""",
                 (student_id, doc_type, f"{doc_type}.pdf", file_path, 'DA_NOP_CHO_KIEM_TRA', now))
    conn.commit()
    conn.close()
    update_overall_status(student_id)
    add_log(session.get('user_id'), 'UPLOAD', student_id, doc_type, f"NÃ¡Â»â„¢p file {doc_type}")
    label = DOC_LABELS.get(doc_type, doc_type)
    return jsonify({'success': True, 'message': f"Ã„ÂÃƒÂ£ nÃ¡Â»â„¢p thÃƒÂ nh cÃƒÂ´ng {label} cho {student['ho_ten']} - LÃ¡Â»â€ºp {student['lop']}."})

@app.route('/api/append-hocba/<int:student_id>', methods=['POST'])
def api_append_hocba(student_id):
    """ThÃƒÂªm trang vÃƒÂ o hÃ¡Â»Âc bÃ¡ÂºÂ¡ Ã„â€˜ÃƒÂ£ nÃ¡Â»â„¢p. NÃ¡ÂºÂ¿u chÃ†Â°a cÃƒÂ³ thÃƒÂ¬ tÃ¡ÂºÂ¡o mÃ¡Â»â€ºi."""
    try:
        student, doc_map = get_student_with_docs(student_id)
        if not student:
            return jsonify({'error': 'KhÃƒÂ´ng tÃƒÂ¬m thÃ¡ÂºÂ¥y hÃ¡Â»Âc sinh.'}), 404
        existing = doc_map.get('HOCBA_6_8', {})
        if existing.get('locked') and existing.get('status') == 'DAT':
            return jsonify({'error': 'HÃ¡Â»Âc bÃ¡ÂºÂ¡ Ã„â€˜ÃƒÂ£ Ã„â€˜Ã†Â°Ã¡Â»Â£c xÃƒÂ¡c nhÃ¡ÂºÂ­n Ã„ÂÃ¡ÂºÂ¡t. LiÃƒÂªn hÃ¡Â»â€¡ giÃƒÂ¡o viÃƒÂªn Ã„â€˜Ã¡Â»Æ’ mÃ¡Â»Å¸ khÃƒÂ³a.'}), 403
        files = request.files.getlist('files')
        if not files:
            return jsonify({'error': 'ChÃ†Â°a chÃ¡Â»Ân file nÃƒÂ o.'}), 400
        max_mb = int(get_setting('max_file_size_mb', '20'))

        # LÃ¡ÂºÂ¥y bytes cÃ¡Â»Â§a file hiÃ¡Â»â€¡n tÃ¡ÂºÂ¡i (local hoÃ¡ÂºÂ·c Drive)
        # NÃ¡ÂºÂ¿u khÃƒÂ´ng tÃ¡ÂºÂ£i Ã„â€˜Ã†Â°Ã¡Â»Â£c (Drive trÃ¡ÂºÂ£ vÃ¡Â»Â HTML, file bÃ¡Â»â€¹ xÃƒÂ³a...) Ã¢â€ â€™ bÃ¡ÂºÂ¯t Ã„â€˜Ã¡ÂºÂ§u mÃ¡Â»â€ºi
        from pdf_utils import append_bytes, _get_bytes
        existing_bytes = None
        existing_path = existing.get('file_path')
        if existing_path:
            try:
                raw, dl_err = _get_bytes(existing_path)
                if not dl_err and raw and len(raw) > 100:
                    # KiÃ¡Â»Æ’m tra Ã„â€˜ÃƒÂ¢y cÃƒÂ³ phÃ¡ÂºÂ£i PDF thÃ¡Â»Â±c khÃƒÂ´ng
                    if raw[:4] == b'%PDF':
                        existing_bytes = raw
            except Exception:
                pass  # KhÃƒÂ´ng tÃ¡ÂºÂ£i Ã„â€˜Ã†Â°Ã¡Â»Â£c Ã¢â€ â€™ coi nhÃ†Â° chÃ†Â°a cÃƒÂ³ file cÃ…Â©

        pdf_bytes, page_count, err = append_bytes(existing_bytes, files, max_mb)
        if err:
            return jsonify({'error': err}), 400

        # LÃ†Â°u vÃƒÂ o Drive hoÃ¡ÂºÂ·c local
        import drive_utils
        if drive_utils.DRIVE_MODE:
            dest_path, err = drive_utils.upload(pdf_bytes, dict(student), 'HOCBA_6_8')
            if err:
                return jsonify({'error': err}), 500
        else:
            folder = get_student_folder(student['lop'], student['ma_hoso'], student['ho_ten_khong_dau'])
            dest_path = os.path.join(folder, 'HOCBA_6_8.pdf')
            bk_dir = os.path.join(os.environ.get('DATA_DIR', os.path.dirname(__file__)), 'backups', student['ma_hoso'])
            if existing_path and os.path.exists(existing_path):
                os.makedirs(bk_dir, exist_ok=True)
                ts = datetime.now().strftime('%Y-%m-%d_%H%M%S')
                shutil.copy2(existing_path, os.path.join(bk_dir, f'backup_HOCBA_6_8_{ts}.pdf'))
            with open(dest_path, 'wb') as f:
                f.write(pdf_bytes)

        conn = get_db()
        now = datetime.now().isoformat()
        conn.execute("DELETE FROM documents WHERE student_id=? AND doc_type='HOCBA_6_8'", (student_id,))
        conn.execute("""INSERT INTO documents (student_id,doc_type,file_name,file_path,status,uploaded_at,note)
                        VALUES (?,?,?,?,?,?,?)""",
                     (student_id, 'HOCBA_6_8', 'HOCBA_6_8.pdf', dest_path,
                      'DA_NOP_CHO_KIEM_TRA', now, f'{page_count} trang'))
        conn.commit()
        conn.close()
        update_overall_status(student_id)
        action = 'APPEND_HOCBA' if existing_path else 'UPLOAD_HOCBA'
        add_log(session.get('user_id'), action, student_id, 'HOCBA_6_8',
                f"{'ThÃƒÂªm' if existing_path else 'TÃ¡ÂºÂ¡o'} hÃ¡Â»Âc bÃ¡ÂºÂ¡: {page_count} trang")
        return jsonify({'success': True,
                        'message': f'Ã„ÂÃƒÂ£ cÃ¡ÂºÂ­p nhÃ¡ÂºÂ­t hÃ¡Â»Âc bÃ¡ÂºÂ¡ ({page_count} trang tÃ¡Â»â€¢ng cÃ¡Â»â„¢ng).',
                        'page_count': page_count})
    except Exception as e:
        return jsonify({'error': f'LÃ¡Â»â€”i mÃƒÂ¡y chÃ¡Â»Â§: {str(e)}'}), 500

@app.route('/api/upload-multi/<int:student_id>/<doc_type>', methods=['POST'])
def api_upload_multi(student_id, doc_type):
    """Upload nhiÃ¡Â»Âu file cÃƒÂ¹ng lÃƒÂºc, gÃ¡Â»â„¢p thÃƒÂ nh 1 PDF (CCCD 2 mÃ¡ÂºÂ·t, v.v.)"""
    doc_type = doc_type.upper()
    if doc_type not in MULTI_FILE_DOCS:
        return jsonify({'error': f'LoÃ¡ÂºÂ¡i tÃƒÂ i liÃ¡Â»â€¡u {doc_type} khÃƒÂ´ng hÃ¡Â»â€” trÃ¡Â»Â£ multi-upload.'}), 400
    student, doc_map = get_student_with_docs(student_id)
    if not student:
        return jsonify({'error': 'KhÃƒÂ´ng tÃƒÂ¬m thÃ¡ÂºÂ¥y hÃ¡Â»Âc sinh.'}), 404
    existing = doc_map.get(doc_type, {})
    if existing.get('locked') and existing.get('status') == 'DAT':
        return jsonify({'error': 'TÃƒÂ i liÃ¡Â»â€¡u Ã„â€˜ÃƒÂ£ Ã„â€˜Ã†Â°Ã¡Â»Â£c xÃƒÂ¡c nhÃ¡ÂºÂ­n Ã„ÂÃ¡ÂºÂ¡t. LiÃƒÂªn hÃ¡Â»â€¡ giÃƒÂ¡o viÃƒÂªn Ã„â€˜Ã¡Â»Æ’ mÃ¡Â»Å¸ khÃƒÂ³a.'}), 403
    files = request.files.getlist('files')
    if not files:
        return jsonify({'error': 'ChÃ†Â°a chÃ¡Â»Ân file nÃƒÂ o.'}), 400
    max_mb = int(get_setting('max_file_size_mb', '20'))

    from pdf_utils import merge_to_bytes
    pdf_bytes, page_count, err = merge_to_bytes(files, max_mb)
    if err:
        return jsonify({'error': err}), 400

    import drive_utils
    if drive_utils.DRIVE_MODE:
        dest_path, err = drive_utils.upload(pdf_bytes, dict(student), doc_type)
        if err:
            return jsonify({'error': err}), 500
    else:
        folder = get_student_folder(student['lop'], student['ma_hoso'], student['ho_ten_khong_dau'])
        dest_path = os.path.join(folder, f'{doc_type}.pdf')
        bk_dir = os.path.join(os.environ.get('DATA_DIR', os.path.dirname(__file__)), 'backups', student['ma_hoso'])
        if os.path.exists(dest_path):
            os.makedirs(bk_dir, exist_ok=True)
            ts = datetime.now().strftime('%Y-%m-%d_%H%M%S')
            shutil.copy2(dest_path, os.path.join(bk_dir, f'backup_{doc_type}_{ts}.pdf'))
        with open(dest_path, 'wb') as f:
            f.write(pdf_bytes)

    conn = get_db()
    now = datetime.now().isoformat()
    conn.execute("DELETE FROM documents WHERE student_id=? AND doc_type=?", (student_id, doc_type))
    conn.execute("""INSERT INTO documents (student_id,doc_type,file_name,file_path,status,uploaded_at,note)
                    VALUES (?,?,?,?,?,?,?)""",
                 (student_id, doc_type, f'{doc_type}.pdf', dest_path,
                  'DA_NOP_CHO_KIEM_TRA', now, f'{page_count} trang'))
    conn.commit()
    conn.close()
    update_overall_status(student_id)
    add_log(session.get('user_id'), f'UPLOAD_MULTI_{doc_type}', student_id, doc_type,
            f"NÃ¡Â»â„¢p {doc_type}: {len(files)} file Ã¢â€ â€™ {page_count} trang")
    return jsonify({'success': True,
                    'message': f'Ã„ÂÃƒÂ£ nÃ¡Â»â„¢p {doc_type} ({len(files)} file, {page_count} trang).',
                    'page_count': page_count})

@app.route('/api/delete-file', methods=['POST'])
def api_delete_file():
    try:
        data = request.get_json(force=True) or {}
        student_id = data.get('student_id')
        doc_type = (data.get('doc_type') or '').upper()
        if not student_id or not doc_type:
            return jsonify({'error': 'DÃ¡Â»Â¯ liÃ¡Â»â€¡u khÃƒÂ´ng hÃ¡Â»Â£p lÃ¡Â»â€¡.'}), 400
        student, doc_map = get_student_with_docs(student_id)
        if not student:
            return jsonify({'error': 'KhÃƒÂ´ng tÃƒÂ¬m thÃ¡ÂºÂ¥y hÃ¡Â»Âc sinh.'}), 404
        doc = doc_map.get(doc_type, {})
        if doc.get('status') == 'DAT' and doc.get('locked'):
            return jsonify({'error': 'File Ã„â€˜ÃƒÂ£ Ã„ÂÃ¡ÂºÂ¡t vÃƒÂ  bÃ¡Â»â€¹ khÃƒÂ³a. LiÃƒÂªn hÃ¡Â»â€¡ giÃƒÂ¡o viÃƒÂªn Ã„â€˜Ã¡Â»Æ’ mÃ¡Â»Å¸ khÃƒÂ³a.'}), 403
        delete_file_to_backup(doc.get('file_path'), student['ma_hoso'], doc_type)
        conn = get_db()
        conn.execute("DELETE FROM documents WHERE student_id=? AND doc_type=?", (student_id, doc_type))
        conn.commit()
        conn.close()
        update_overall_status(student_id)
        add_log(session.get('user_id'), 'DELETE', student_id, doc_type, f"XÃƒÂ³a file {doc_type}")
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': f'LÃ¡Â»â€”i mÃƒÂ¡y chÃ¡Â»Â§: {str(e)}'}), 500

@app.route('/api/check-file', methods=['POST'])
@login_required
def api_check_file():
    data = request.get_json()
    student_id = data.get('student_id')
    doc_type = data.get('doc_type', '').upper()
    new_status = data.get('status', '')
    note = data.get('note', '')
    valid_statuses = ['DAT', 'FILE_MO', 'SAI_FILE', 'THIEU_TRANG', 'CAN_NOP_LAI']
    if new_status not in valid_statuses:
        return jsonify({'error': 'TrÃ¡ÂºÂ¡ng thÃƒÂ¡i khÃƒÂ´ng hÃ¡Â»Â£p lÃ¡Â»â€¡.'}), 400
    conn = get_db()
    locked = 1 if new_status == 'DAT' else 0
    now = datetime.now().isoformat()
    conn.execute("""UPDATE documents SET status=?, note=?, checked_by=?, checked_at=?, locked=?
                    WHERE student_id=? AND doc_type=?""",
                 (new_status, note, session.get('username'), now, locked, student_id, doc_type))
    conn.commit()
    conn.close()
    update_overall_status(student_id)
    add_log(session.get('user_id'), 'CHECK', student_id, doc_type, f"Ã„ÂÃƒÂ¡nh dÃ¡ÂºÂ¥u {new_status}: {note}")
    return jsonify({'success': True})

@app.route('/api/unlock-file', methods=['POST'])
@login_required
def api_unlock_file():
    data = request.get_json()
    student_id = data.get('student_id')
    doc_type = data.get('doc_type', '').upper()
    conn = get_db()
    conn.execute("UPDATE documents SET locked=0, status='CAN_NOP_LAI' WHERE student_id=? AND doc_type=?",
                 (student_id, doc_type))
    conn.commit()
    conn.close()
    update_overall_status(student_id)
    add_log(session.get('user_id'), 'UNLOCK', student_id, doc_type, "MÃ¡Â»Å¸ khÃƒÂ³a cho nÃ¡Â»â„¢p lÃ¡ÂºÂ¡i")
    return jsonify({'success': True})

@app.route('/api/merge-hocba-parts/<int:student_id>', methods=['POST'])
def api_merge_hocba_parts(student_id):
    """NÃ¡Â»â€˜i nhiÃ¡Â»Âu file PDF thÃƒÂ nh HOCBA_6_8 Ã¢â‚¬â€ hÃ¡Â»Âc sinh cÃƒÂ³ thÃ¡Â»Æ’ tÃ¡Â»Â± thÃ¡Â»Â±c hiÃ¡Â»â€¡n"""
    try:
        from pdf_utils import merge_multiple_pdfs
        student, doc_map = get_student_with_docs(student_id)
        if not student:
            return jsonify({'error': 'KhÃƒÂ´ng tÃƒÂ¬m thÃ¡ÂºÂ¥y hÃ¡Â»Âc sinh.'}), 404
        files = request.files.getlist('files')
        if not files or len(files) == 0:
            return jsonify({'error': 'ChÃ†Â°a chÃ¡Â»Ân file nÃƒÂ o.'}), 400
        max_mb = int(get_setting('max_file_size_mb', '20'))
        folder = get_student_folder(student['lop'], student['ma_hoso'], student['ho_ten_khong_dau'])
        dest = os.path.join(folder, 'HOCBA_6_8.pdf')
        bk_dir = os.path.join(os.environ.get('DATA_DIR', os.path.dirname(__file__)), 'backups', student['ma_hoso'])
        dest_path, err = merge_multiple_pdfs(files, dest, bk_dir, 'HOCBA_6_8', max_mb)
        if err:
            return jsonify({'error': err}), 400
        conn = get_db()
        now = datetime.now().isoformat()
        conn.execute("DELETE FROM documents WHERE student_id=? AND doc_type='HOCBA_6_8'", (student_id,))
        conn.execute("INSERT INTO documents (student_id,doc_type,file_name,file_path,status,uploaded_at) VALUES (?,?,?,?,?,?)",
                     (student_id, 'HOCBA_6_8', 'HOCBA_6_8.pdf', dest_path, 'DA_NOP_CHO_KIEM_TRA', now))
        conn.commit()
        conn.close()
        update_overall_status(student_id)
        add_log(None, 'UPLOAD_MERGE', student_id, 'HOCBA_6_8', f"NÃ¡Â»â€˜i {len(files)} file thÃƒÂ nh HOCBA_6_8")
        return jsonify({'success': True, 'message': f'Ã„ÂÃƒÂ£ nÃ¡Â»â€˜i {len(files)} file thÃƒÂ nh HÃ¡Â»Âc bÃ¡ÂºÂ¡ 6-8 thÃƒÂ nh cÃƒÂ´ng!'})
    except Exception as e:
        return jsonify({'error': f'LÃ¡Â»â€”i: {str(e)}'}), 500

@app.route('/api/merge-transcript/<int:student_id>', methods=['POST'])
def api_merge_transcript(student_id):
    student, doc_map = get_student_with_docs(student_id)
    if not student:
        return jsonify({'error': 'KhÃƒÂ´ng tÃƒÂ¬m thÃ¡ÂºÂ¥y hÃ¡Â»Âc sinh.'}), 404
    folder = get_student_folder(student['lop'], student['ma_hoso'], student['ho_ten_khong_dau'])
    dest, err = merge_transcripts(folder, student['ma_hoso'], doc_map)
    if err:
        return jsonify({'error': err}), 400
    conn = get_db()
    now = datetime.now().isoformat()
    conn.execute("DELETE FROM documents WHERE student_id=? AND doc_type='HOCBA'", (student_id,))
    conn.execute("INSERT INTO documents (student_id,doc_type,file_name,file_path,status,uploaded_at) VALUES (?,?,?,?,?,?)",
                 (student_id, 'HOCBA', 'HOCBA.pdf', dest, 'DA_NOP_CHO_KIEM_TRA', now))
    conn.commit()
    conn.close()
    update_overall_status(student_id)
    add_log(session.get('user_id'), 'MERGE', student_id, 'HOCBA', "NÃ¡Â»â€˜i hÃ¡Â»Âc bÃ¡ÂºÂ¡ thÃƒÂ nh cÃƒÂ´ng")
    return jsonify({'success': True, 'message': f"Ã„ÂÃƒÂ£ nÃ¡Â»â€˜i hÃ¡Â»Âc bÃ¡ÂºÂ¡ cho {student['ho_ten']}."})

@app.route('/api/merge-transcript-class/<class_name>', methods=['POST'])
@login_required
def api_merge_class(class_name):
    conn = get_db()
    students = conn.execute("SELECT * FROM students WHERE lop=?", (class_name,)).fetchall()
    conn.close()
    ok, fail = [], []
    for s in students:
        s = dict(s)
        doc_map = get_student_docs(s['id'])
        folder = get_student_folder(s['lop'], s['ma_hoso'], s['ho_ten_khong_dau'])
        dest, err = merge_transcripts(folder, s['ma_hoso'], doc_map)
        if err:
            fail.append({'ho_ten': s['ho_ten'], 'ly_do': err})
        else:
            conn2 = get_db()
            now = datetime.now().isoformat()
            conn2.execute("DELETE FROM documents WHERE student_id=? AND doc_type='HOCBA'", (s['id'],))
            conn2.execute("INSERT INTO documents (student_id,doc_type,file_name,file_path,status,uploaded_at) VALUES (?,?,?,?,?,?)",
                         (s['id'], 'HOCBA', 'HOCBA.pdf', dest, 'DA_NOP_CHO_KIEM_TRA', now))
            conn2.commit()
            conn2.close()
            update_overall_status(s['id'])
            ok.append(s['ho_ten'])
    return jsonify({'success_count': len(ok), 'fail_count': len(fail), 'failures': fail})

@app.route('/api/merge-transcript-all', methods=['POST'])
@login_required
@admin_required
def api_merge_all():
    conn = get_db()
    students = conn.execute("SELECT * FROM students").fetchall()
    conn.close()
    ok, fail = [], []
    for s in students:
        s = dict(s)
        doc_map = get_student_docs(s['id'])
        folder = get_student_folder(s['lop'], s['ma_hoso'], s['ho_ten_khong_dau'])
        dest, err = merge_transcripts(folder, s['ma_hoso'], doc_map)
        if err:
            fail.append({'ho_ten': s['ho_ten'], 'lop': s['lop'], 'ly_do': err})
        else:
            conn2 = get_db()
            now = datetime.now().isoformat()
            conn2.execute("DELETE FROM documents WHERE student_id=? AND doc_type='HOCBA'", (s['id'],))
            conn2.execute("INSERT INTO documents (student_id,doc_type,file_name,file_path,status,uploaded_at) VALUES (?,?,?,?,?,?)",
                         (s['id'], 'HOCBA', 'HOCBA.pdf', dest, 'DA_NOP_CHO_KIEM_TRA', now))
            conn2.commit()
            conn2.close()
            update_overall_status(s['id'])
            ok.append(s['ho_ten'])
    return jsonify({'success_count': len(ok), 'fail_count': len(fail), 'failures': fail})

@app.route('/api/export-excel')
@login_required
def api_export_excel():
    class_name = request.args.get('class', 'ALL')
    conn = get_db()
    if class_name == 'ALL':
        rows = conn.execute("SELECT * FROM students ORDER BY lop, stt").fetchall()
    else:
        rows = conn.execute("SELECT * FROM students WHERE lop=? ORDER BY stt", (class_name,)).fetchall()
    conn.close()
    students = enrich_students(rows)
    buf = export_excel(students, class_name)
    fname = f"HoSo_{class_name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/download-template')
@login_required
@admin_required
def download_template():
    """TÃ¡ÂºÂ¡o vÃƒÂ  tÃ¡ÂºÂ£i file Excel mÃ¡ÂºÂ«u Ã„â€˜Ã¡Â»Æ’ import danh sÃƒÂ¡ch hÃ¡Â»Âc sinh"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Danh sÃƒÂ¡ch hÃ¡Â»Âc sinh"
    from openpyxl.styles import Font, PatternFill, Alignment
    # Header
    headers = ['ma_hoso', 'lop', 'stt', 'ho_ten', 'ngay_sinh', 'ghi_chu']
    header_labels = ['MÃƒÂ£ hÃ¡Â»â€œ sÃ†Â¡', 'LÃ¡Â»â€ºp', 'STT', 'HÃ¡Â»Â tÃƒÂªn', 'NgÃƒÂ y sinh', 'Ghi chÃƒÂº']
    for col, (key, label) in enumerate(zip(headers, header_labels), 1):
        cell = ws.cell(row=1, column=col, value=key)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='1565C0')
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[cell.column_letter].width = 18
        ws.cell(row=2, column=col, value=label).font = Font(italic=True, color='555555')
    # DÃ¡Â»Â¯ liÃ¡Â»â€¡u mÃ¡ÂºÂ«u
    samples = [
        ['9B1_001', '9B1', 1, 'NguyÃ¡Â»â€¦n VÃ„Æ’n An', '12/05/2011', ''],
        ['9B1_002', '9B1', 2, 'TrÃ¡ÂºÂ§n ThÃ¡Â»â€¹ BÃƒÂ¬nh', '03/07/2011', ''],
        ['9B2_001', '9B2', 1, 'LÃƒÂª Minh ChÃƒÂ¢u',  '25/09/2011', ''],
    ]
    for i, row in enumerate(samples, 3):
        for j, val in enumerate(row, 1):
            ws.cell(row=i, column=j, value=val)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name='danh_sach_hoc_sinh_mau.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/api/download-student-zip/<int:student_id>')
@login_required
def api_download_student_zip(student_id):
    student, doc_map = get_student_with_docs(student_id)
    if not student:
        return jsonify({'error': 'KhÃƒÂ´ng tÃƒÂ¬m thÃ¡ÂºÂ¥y'}), 404
    buf = create_student_zip(student, doc_map)
    fname = f"{student['ma_hoso']}_{student['ho_ten_khong_dau']}.zip"
    return send_file(buf, as_attachment=True, download_name=fname, mimetype='application/zip')

@app.route('/api/download-class-zip/<class_name>')
@login_required
def api_download_class_zip(class_name):
    conn = get_db()
    rows = conn.execute("SELECT * FROM students WHERE lop=? ORDER BY stt", (class_name,)).fetchall()
    conn.close()
    students = enrich_students(rows)
    buf = create_class_zip(class_name, students)
    fname = f"HoSo_Lop_{class_name}_{datetime.now().strftime('%Y%m%d')}.zip"
    return send_file(buf, as_attachment=True, download_name=fname, mimetype='application/zip')

@app.route('/api/download-all-zip')
@login_required
@admin_required
def api_download_all_zip():
    conn = get_db()
    rows = conn.execute("SELECT * FROM students ORDER BY lop, stt").fetchall()
    conn.close()
    students = enrich_students(rows)
    buf = create_all_zip(students)
    fname = f"HoSo_ToanKhoi_{datetime.now().strftime('%Y%m%d')}.zip"
    return send_file(buf, as_attachment=True, download_name=fname, mimetype='application/zip')

@app.route('/api/import-students', methods=['POST'])
@login_required
@admin_required
def api_import_students():
    file = request.files.get('file')
    if not file:
        return jsonify({'error': 'ChÃ†Â°a chÃ¡Â»Ân file.'}), 400
    ext = file.filename.rsplit('.', 1)[-1].lower()
    rows_data = []
    try:
        if ext in ('xlsx', 'xls'):
            wb = openpyxl.load_workbook(file, read_only=True)
            ws = wb.active
            headers = [str(c.value).strip() if c.value else '' for c in next(ws.iter_rows())]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(v for v in row):
                    def _cell_str(v):
                        if v is None:
                            return ''
                        # openpyxl trÃ¡ÂºÂ£ ngÃƒÂ y dÃ†Â°Ã¡Â»â€ºi dÃ¡ÂºÂ¡ng datetime
                        try:
                            from datetime import datetime as _dt
                            if isinstance(v, _dt):
                                return v.strftime('%d/%m/%Y')
                        except Exception:
                            pass
                        return str(v).strip()
                    rows_data.append(dict(zip(headers, [_cell_str(v) for v in row])))
        elif ext == 'csv':
            import csv, io as _io
            content = file.read().decode('utf-8-sig')
            reader = csv.DictReader(_io.StringIO(content))
            for row in reader:
                rows_data.append({k.strip(): v.strip() for k, v in row.items()})
        else:
            return jsonify({'error': 'ChÃ¡Â»â€° hÃ¡Â»â€” trÃ¡Â»Â£ Excel (.xlsx) hoÃ¡ÂºÂ·c CSV.'}), 400
    except Exception as e:
        return jsonify({'error': f'LÃ¡Â»â€”i Ã„â€˜Ã¡Â»Âc file: {str(e)}'}), 400

    conn = get_db()
    overwrite = request.form.get('overwrite', '1') == '1'  # mÃ¡ÂºÂ·c Ã„â€˜Ã¡Â»â€¹nh ghi Ã„â€˜ÃƒÂ¨ (hÃƒÂ nh vi cÃ…Â©)
    # LÃ¡ÂºÂ¥y danh sÃƒÂ¡ch ma_hoso Ã„â€˜ÃƒÂ£ tÃ¡Â»â€œn tÃ¡ÂºÂ¡i
    existing_ma = {r['ma_hoso'] for r in conn.execute('SELECT ma_hoso FROM students').fetchall()}
    count, skipped, errors = 0, 0, []
    for i, r in enumerate(rows_data, 2):
        ma_hoso = r.get('ma_hoso', '').strip()
        lop = r.get('lop', '').strip()
        stt_raw = r.get('stt', '').strip()
        try:
            stt = str(int(float(stt_raw))) if stt_raw else ''
        except (ValueError, TypeError):
            stt = stt_raw
        ho_ten = r.get('ho_ten', '').strip()
        ngay_sinh_raw = r.get('ngay_sinh', '').strip()
        if ngay_sinh_raw:
            import re as _re
            m = _re.match(r'^(\d{4})-(\d{1,2})-(\d{1,2})', ngay_sinh_raw)
            if m:
                ngay_sinh = f"{m.group(3).zfill(2)}/{m.group(2).zfill(2)}/{m.group(1)}"
            else:
                ngay_sinh = ngay_sinh_raw
        else:
            ngay_sinh = ''
        if not all([ma_hoso, lop, ho_ten]):
            errors.append(f"DÃƒÂ²ng {i}: thiÃ¡ÂºÂ¿u thÃƒÂ´ng tin bÃ¡ÂºÂ¯t buÃ¡Â»â„¢c.")
            continue
        # KiÃ¡Â»Æ’m tra trÃƒÂ¹ng
        if ma_hoso in existing_ma and not overwrite:
            skipped += 1
            continue
        ho_ten_khong_dau = to_ascii(ho_ten)
        now = datetime.now().isoformat()
        try:
            conn.execute("""INSERT OR REPLACE INTO students
                            (ma_hoso,lop,stt,ho_ten,ho_ten_khong_dau,ngay_sinh,status_overall,created_at,updated_at)
                            VALUES (?,?,?,?,?,?,?,?,?)""",
                         (ma_hoso, lop, stt, ho_ten, ho_ten_khong_dau, ngay_sinh, 'CHUA_NOP', now, now))
            count += 1
        except Exception as e:
            errors.append(f"DÃƒÂ²ng {i}: {str(e)}")
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'imported': count, 'skipped': skipped, 'errors': errors})

@app.route('/api/preview-import', methods=['POST'])
@login_required
@admin_required
def api_preview_import():
    """PhÃƒÂ¢n tÃƒÂ­ch file Excel: trÃ¡ÂºÂ£ vÃ¡Â»Â danh sÃƒÂ¡ch mÃ¡Â»â€ºi / trÃƒÂ¹ng (dÃ¡Â»Â±a vÃƒÂ o ma_hoso)."""
    file = request.files.get('file')
    if not file:
        return jsonify({'error': 'ChÃ†Â°a chÃ¡Â»Ân file.'}), 400
    ext = file.filename.rsplit('.', 1)[-1].lower()
    rows_data = []
    try:
        if ext in ('xlsx', 'xls'):
            wb = openpyxl.load_workbook(file, read_only=True)
            ws = wb.active
            headers = [str(c.value).strip() if c.value else '' for c in next(ws.iter_rows())]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(v for v in row):
                    def _cs(v):
                        if v is None: return ''
                        try:
                            from datetime import datetime as _dt
                            if isinstance(v, _dt): return v.strftime('%d/%m/%Y')
                        except Exception: pass
                        return str(v).strip()
                    rows_data.append(dict(zip(headers, [_cs(v) for v in row])))
        elif ext == 'csv':
            import csv, io as _io
            content = file.read().decode('utf-8-sig')
            reader = csv.DictReader(_io.StringIO(content))
            for row in reader:
                rows_data.append({k.strip(): v.strip() for k, v in row.items()})
        else:
            return jsonify({'error': 'ChÃ¡Â»â€° hÃ¡Â»â€” trÃ¡Â»Â£ .xlsx hoÃ¡ÂºÂ·c .csv'}), 400
    except Exception as e:
        return jsonify({'error': f'LÃ¡Â»â€”i Ã„â€˜Ã¡Â»Âc file: {str(e)}'}), 400

    conn = get_db()
    existing_ma = {r['ma_hoso']: dict(r) for r in conn.execute('SELECT ma_hoso, ho_ten, lop FROM students').fetchall()}
    conn.close()

    new_list, dup_list, err_list = [], [], []
    for i, r in enumerate(rows_data, 2):
        ma = r.get('ma_hoso', '').strip()
        ho_ten = r.get('ho_ten', '').strip()
        lop = r.get('lop', '').strip()
        if not all([ma, ho_ten, lop]):
            err_list.append({'row': i, 'reason': 'ThiÃ¡ÂºÂ¿u thÃƒÂ´ng tin bÃ¡ÂºÂ¯t buÃ¡Â»â„¢c'})
            continue
        if ma in existing_ma:
            dup_list.append({'ma_hoso': ma, 'ho_ten_moi': ho_ten, 'lop_moi': lop,
                             'ho_ten_cu': existing_ma[ma]['ho_ten'], 'lop_cu': existing_ma[ma]['lop']})
        else:
            new_list.append({'ma_hoso': ma, 'ho_ten': ho_ten, 'lop': lop})
    return jsonify({'success': True, 'new': new_list, 'duplicates': dup_list, 'errors': err_list,
                    'total': len(rows_data)})

@app.route('/api/fix-data', methods=['POST'])
@login_required
@admin_required
def api_fix_data():
    """DÃ¡Â»Ân dÃ¡Â»Â¯ liÃ¡Â»â€¡u cÃ…Â©: chuÃ¡ÂºÂ©n hÃƒÂ³a stt (1.0Ã¢â€ â€™1) vÃƒÂ  ngÃƒÂ y sinh (YYYY-MM-DDÃ¢â€ â€™dd/mm/YYYY)."""
    import re
    conn = get_db()
    rows = conn.execute("SELECT id, stt, ngay_sinh FROM students").fetchall()
    fixed = 0
    for row in rows:
        new_stt = row['stt']
        new_date = row['ngay_sinh']
        changed = False
        # Fix STT: "1.0" Ã¢â€ â€™ "1"
        if new_stt:
            try:
                clean = str(int(float(new_stt.strip())))
                if clean != new_stt.strip():
                    new_stt = clean
                    changed = True
            except (ValueError, AttributeError):
                pass
        # Fix ngÃƒÂ y sinh: YYYY-MM-DD... Ã¢â€ â€™ dd/mm/YYYY
        if new_date:
            m = re.match(r'^(\d{4})-(\d{1,2})-(\d{1,2})', new_date.strip())
            if m:
                new_date = f"{m.group(3).zfill(2)}/{m.group(2).zfill(2)}/{m.group(1)}"
                changed = True
        if changed:
            conn.execute("UPDATE students SET stt=?, ngay_sinh=? WHERE id=?",
                         (new_stt, new_date, row['id']))
            fixed += 1
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'fixed': fixed, 'message': f'Ã„ÂÃƒÂ£ chuÃ¡ÂºÂ©n hÃƒÂ³a {fixed} bÃ¡ÂºÂ£n ghi.'})

@app.route('/api/settings', methods=['POST'])
@login_required
@admin_required
def api_settings():
    data = request.get_json()
    for key in ['phase', 'max_file_size_mb']:
        if key in data:
            set_setting(key, str(data[key]))
    return jsonify({'success': True})

@app.route('/admin/import')
@login_required
@admin_required
def admin_import():
    return render_template('admin_import.html')

@app.route('/admin/settings')
@login_required
@admin_required
def admin_settings():
    phase = get_setting('phase', '1')
    max_mb = get_setting('max_file_size_mb', '5')
    conn = get_db()
    teachers = conn.execute("SELECT * FROM users WHERE role IN ('teacher','admin') ORDER BY full_name").fetchall()
    classes = [r['lop'] for r in conn.execute("SELECT DISTINCT lop FROM students ORDER BY lop").fetchall()]
    conn.close()
    # TÃƒÂ­nh dung lÃ†Â°Ã¡Â»Â£ng Ã„â€˜Ã¡Â»Æ’ monitor quota PythonAnywhere
    disk_info = {'uploads_mb': 0, 'backups_mb': 0, 'db_mb': 0, 'total_mb': 0, 'pct': 0, 'quota_mb': 512}
    try:
        from file_utils import UPLOAD_FOLDER, BACKUP_FOLDER
        from database import DB_PATH as _db_path
        def _folder_mb(path):
            total = 0
            if os.path.exists(path):
                for root, _, files in os.walk(path):
                    for fn in files:
                        try: total += os.path.getsize(os.path.join(root, fn))
                        except OSError: pass
            return round(total / 1024 / 1024, 1)
        disk_info['uploads_mb'] = _folder_mb(UPLOAD_FOLDER)
        disk_info['backups_mb'] = _folder_mb(BACKUP_FOLDER)
        disk_info['db_mb']      = round(os.path.getsize(_db_path) / 1024 / 1024, 2) if os.path.exists(_db_path) else 0
        disk_info['total_mb']   = disk_info['uploads_mb'] + disk_info['backups_mb'] + disk_info['db_mb']
        disk_info['pct']        = min(round(disk_info['total_mb'] / disk_info['quota_mb'] * 100, 1), 100)
    except Exception:
        pass
    return render_template('admin_settings.html', phase=phase, max_mb=max_mb,
                           teachers=teachers, classes=classes, disk_info=disk_info)

@app.route('/api/add-teacher', methods=['POST'])
@login_required
@admin_required
def api_add_teacher():
    data = request.get_json()
    username = data.get('username', '').strip()
    password = data.get('password', '').strip()
    full_name = data.get('full_name', '').strip()
    assigned_class = data.get('assigned_class', '').strip()
    if not username or not password:
        return jsonify({'error': 'ThiÃ¡ÂºÂ¿u tÃƒÂªn Ã„â€˜Ã„Æ’ng nhÃ¡ÂºÂ­p hoÃ¡ÂºÂ·c mÃ¡ÂºÂ­t khÃ¡ÂºÂ©u.'}), 400
    conn = get_db()
    try:
        conn.execute("INSERT INTO users (username,password_hash,role,full_name,assigned_class,created_at) VALUES (?,?,?,?,?,?)",
                     (username, generate_password_hash(password), 'teacher', full_name, assigned_class, datetime.now().isoformat()))
        conn.commit()
    except Exception as e:
        conn.close()
        return jsonify({'error': f'LÃ¡Â»â€”i: {str(e)}'}), 400
    conn.close()
    return jsonify({'success': True})

@app.route('/api/change-password', methods=['POST'])
@login_required
def api_change_password():
    data = request.get_json()
    old_pw = data.get('old_password', '')
    new_pw = data.get('new_password', '').strip()
    if not new_pw or len(new_pw) < 6:
        return jsonify({'error': 'MÃ¡ÂºÂ­t khÃ¡ÂºÂ©u mÃ¡Â»â€ºi phÃ¡ÂºÂ£i cÃƒÂ³ ÃƒÂ­t nhÃ¡ÂºÂ¥t 6 kÃƒÂ½ tÃ¡Â»Â±.'}), 400
    conn = get_db()
    user = conn.execute("SELECT * FROM users WHERE id=?", (session['user_id'],)).fetchone()
    if not check_password_hash(user['password_hash'], old_pw):
        conn.close()
        return jsonify({'error': 'MÃ¡ÂºÂ­t khÃ¡ÂºÂ©u cÃ…Â© khÃƒÂ´ng Ã„â€˜ÃƒÂºng.'}), 400
    conn.execute("UPDATE users SET password_hash=? WHERE id=?",
                 (generate_password_hash(new_pw), session['user_id']))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/edit-student', methods=['POST'])
@login_required
def api_edit_student():
    """ChÃ¡Â»â€°nh sÃ¡Â»Â­a thÃƒÂ´ng tin hÃ¡Â»Âc sinh. GiÃƒÂ¡o viÃƒÂªn/admin Ã„â€˜Ã†Â°Ã¡Â»Â£c phÃƒÂ©p."""
    data = request.get_json()
    student_id = data.get('student_id')
    if not student_id:
        return jsonify({'error': 'ThiÃ¡ÂºÂ¿u student_id.'}), 400
    # KiÃ¡Â»Æ’m tra quyÃ¡Â»Ân: chÃ¡Â»â€° teacher/admin mÃ¡Â»â€ºi Ã„â€˜Ã†Â°Ã¡Â»Â£c sÃ¡Â»Â­a
    if session.get('role') not in ('teacher', 'admin'):
        return jsonify({'error': 'BÃ¡ÂºÂ¡n khÃƒÂ´ng cÃƒÂ³ quyÃ¡Â»Ân chÃ¡Â»â€°nh sÃ¡Â»Â­a.'}), 403
    ho_ten = (data.get('ho_ten') or '').strip()
    ngay_sinh = (data.get('ngay_sinh') or '').strip()
    lop = (data.get('lop') or '').strip()
    stt = (data.get('stt') or '').strip()
    ghi_chu = (data.get('ghi_chu') or '').strip()
    cccd = (data.get('cccd') or '').strip()
    ma_dinh_danh_gd = (data.get('ma_dinh_danh_gd') or '').strip()
    if not ho_ten:
        return jsonify({'error': 'HÃ¡Â»Â tÃƒÂªn khÃƒÂ´ng Ã„â€˜Ã†Â°Ã¡Â»Â£c Ã„â€˜Ã¡Â»Æ’ trÃ¡Â»â€˜ng.'}), 400
    ho_ten_khong_dau = to_ascii(ho_ten)
    try:
        conn = get_db()
        conn.execute("""UPDATE students SET ho_ten=?, ho_ten_khong_dau=?, ngay_sinh=?,
                        lop=?, stt=?, note=?, cccd=?, ma_dinh_danh_gd=?, updated_at=? WHERE id=?""",
                     (ho_ten, ho_ten_khong_dau, ngay_sinh, lop, stt, ghi_chu,
                      cccd or None, ma_dinh_danh_gd or None,
                      datetime.now().isoformat(), student_id))
        conn.commit()
        conn.close()
        add_log(None, 'edit_student', student_id=student_id, detail=f"GV sÃ¡Â»Â­a: {ho_ten}")
    except Exception as e:
        return jsonify({'error': f'LÃ¡Â»â€”i cÃ¡ÂºÂ­p nhÃ¡ÂºÂ­t: {str(e)}'}), 500
    return jsonify({'success': True, 'message': 'Ã„ÂÃƒÂ£ cÃ¡ÂºÂ­p nhÃ¡ÂºÂ­t thÃƒÂ´ng tin hÃ¡Â»Âc sinh.'})

@app.route('/api/delete-student', methods=['POST'])
@login_required
def api_delete_student():
    """XÃƒÂ³a hÃ¡Â»Âc sinh vÃƒÂ  toÃƒÂ n bÃ¡Â»â„¢ file Ã„â€˜i kÃƒÂ¨m (chÃ¡Â»â€° teacher/admin)."""
    if session.get('role') not in ('teacher', 'admin'):
        return jsonify({'error': 'BÃ¡ÂºÂ¡n khÃƒÂ´ng cÃƒÂ³ quyÃ¡Â»Ân xÃƒÂ³a hÃ¡Â»Âc sinh.'}), 403
    data = request.get_json(force=True) or {}
    student_id = data.get('student_id')
    if not student_id:
        return jsonify({'error': 'ThiÃ¡ÂºÂ¿u student_id.'}), 400
    conn = get_db()
    row = conn.execute("SELECT * FROM students WHERE id=?", (student_id,)).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': 'KhÃƒÂ´ng tÃƒÂ¬m thÃ¡ÂºÂ¥y hÃ¡Â»Âc sinh.'}), 404
    # XÃƒÂ³a document records + backup file_path (tuÃ¡Â»Â³ chÃ¡Â»Ân Ã¢â‚¬â€œ chÃ¡Â»â€° xÃƒÂ³a record, giÃ¡Â»Â¯ file)
    conn.execute("DELETE FROM documents WHERE student_id=?", (student_id,))
    conn.execute("DELETE FROM students WHERE id=?", (student_id,))
    conn.commit()
    conn.close()
    add_log(session.get('user_id'), 'DELETE_STUDENT', student_id, None,
            f"XÃƒÂ³a HS: {row['ho_ten']} - LÃ¡Â»â€ºp {row['lop']}")
    return jsonify({'success': True, 'message': f"Ã„ÂÃƒÂ£ xÃƒÂ³a hÃ¡Â»Âc sinh {row['ho_ten']}."})

@app.route('/api/add-student', methods=['POST'])
@login_required
def api_add_student():
    """ThÃƒÂªm hÃ¡Â»Âc sinh thÃ¡Â»Â§ cÃƒÂ´ng (chÃ¡Â»â€° teacher/admin)."""
    if session.get('role') not in ('teacher', 'admin'):
        return jsonify({'error': 'BÃ¡ÂºÂ¡n khÃƒÂ´ng cÃƒÂ³ quyÃ¡Â»Ân thÃƒÂªm hÃ¡Â»Âc sinh.'}), 403
    data = request.get_json(force=True) or {}
    ho_ten = (data.get('ho_ten') or '').strip()
    lop = (data.get('lop') or '').strip()
    stt = (data.get('stt') or '1').strip()
    ngay_sinh = (data.get('ngay_sinh') or '').strip()
    if not ho_ten or not lop:
        return jsonify({'error': 'HÃ¡Â»Â tÃƒÂªn vÃƒÂ  lÃ¡Â»â€ºp khÃƒÂ´ng Ã„â€˜Ã†Â°Ã¡Â»Â£c Ã„â€˜Ã¡Â»Æ’ trÃ¡Â»â€˜ng.'}), 400
    ho_ten_khong_dau = to_ascii(ho_ten)
    now = datetime.now().isoformat()
    # TÃ¡ÂºÂ¡o mÃƒÂ£ hÃ¡Â»â€œ sÃ†Â¡ tÃ¡Â»Â± Ã„â€˜Ã¡Â»â„¢ng: LOP_STT_HOTENKD
    ma_hoso = f"{to_ascii(lop)}_{stt}_{ho_ten_khong_dau}"[:60]
    try:
        conn = get_db()
        conn.execute("""INSERT INTO students
                        (ma_hoso,lop,stt,ho_ten,ho_ten_khong_dau,ngay_sinh,status_overall,created_at,updated_at)
                        VALUES (?,?,?,?,?,?,?,?,?)""",
                     (ma_hoso, lop, stt, ho_ten, ho_ten_khong_dau, ngay_sinh,
                      'CHUA_NOP', now, now))
        conn.commit()
        new_id = conn.execute("SELECT id FROM students WHERE ma_hoso=?", (ma_hoso,)).fetchone()['id']
        conn.close()
        add_log(session.get('user_id'), 'ADD_STUDENT', new_id, None,
                f"ThÃƒÂªm HS: {ho_ten} - LÃ¡Â»â€ºp {lop}")
        return jsonify({'success': True, 'message': f"Ã„ÂÃƒÂ£ thÃƒÂªm hÃ¡Â»Âc sinh {ho_ten}.", 'id': new_id})
    except Exception as e:
        return jsonify({'error': f'LÃ¡Â»â€”i: {str(e)}'}), 500

@app.route('/view-file/<int:student_id>/<doc_type>')
def view_file(student_id, doc_type):
    """Xem file PDF Ã¢â‚¬â€ redirect sang Drive URL nÃ¡ÂºÂ¿u lÃƒÂ  Drive file."""
    student, doc_map = get_student_with_docs(student_id)
    doc = doc_map.get(doc_type.upper(), {})
    fp  = doc.get('file_path')
    if not fp:
        return '<p style="font-family:sans-serif;padding:20px">Ã¢ÂÅ’ File chÃ†Â°a Ã„â€˜Ã†Â°Ã¡Â»Â£c nÃ¡Â»â„¢p. Vui lÃƒÂ²ng nÃ¡Â»â„¢p lÃ¡ÂºÂ¡i.</p>', 404
    import drive_utils
    if drive_utils.is_drive(fp):
        return redirect(drive_utils.view_url(fp))
    if not os.path.exists(fp):
        return '<p style="font-family:sans-serif;padding:20px">Ã¢ÂÅ’ File khÃƒÂ´ng tÃ¡Â»â€œn tÃ¡ÂºÂ¡i. Vui lÃƒÂ²ng nÃ¡Â»â„¢p lÃ¡ÂºÂ¡i.</p>', 404
    return send_file(fp, mimetype='application/pdf')

# ===== TSDC INTEGRATION =====
_tsdc_cache = {'data': None, 'ts': 0, 'fetching': False}
_TSDC_TTL   = 300  # cache 5 phut

# JavaScript extract (raw string de tranh loi escape Python/JS)
_TSDC_JS = r"""
() => {
    var rows = Array.from(document.querySelectorAll('tbody tr'));
    var students = [];
    function isMaHS(t) {
        return t.startsWith('HS') && !t.startsWith('HSO') &&
               t.length > 5 && t.charCodeAt(2) >= 48 && t.charCodeAt(2) <= 57;
    }
    function isDate(t) { return t.length===10&&t.charAt(2)==='/'&&t.charAt(5)==='/'; }
    function allDigits(t) {
        for(var i=0;i<t.length;i++) if(t.charCodeAt(i)<48||t.charCodeAt(i)>57) return false;
        return true;
    }
    function isIdOrCCCD(t) {
        if(t.length<9||t.length>12||!allDigits(t)) return false;
        if(t.length===10&&t.charAt(0)==='0') return false;
        return true;
    }
    function isLop(t) {
        if(t.length<2||t.length>5) return false;
        if(t.charAt(0)==='9'){var c=t.charCodeAt(1);return c>=65&&c<=90;}
        if(t.charAt(0)==='1'&&t.charAt(1)==='0'&&t.length>2){var c=t.charCodeAt(2);return c>=65&&c<=90;}
        return false;
    }
    function isNVSchool(t) {
        return t.indexOf('THPT')!==-1||t.indexOf('PTDT')!==-1||
               t.indexOf('Lien Viet')!==-1||t.indexOf('Lien Viet')!==-1;
    }
    rows.forEach(function(row) {
        var cells=Array.from(row.querySelectorAll('td')).map(function(c){return c.innerText.trim();});
        if(!cells.some(function(t){return isMaHS(t);})) return;
        var s={id:'',maHocSinh:'',hoTen:'',trangThai:'',ngaySinh:'',gioiTinh:'',lop:'',nv1:'',nv2:'',nv3:''};
        var afterLop=false,nvs=[];
        cells.forEach(function(t) {
            if(!t) return;
            if(isMaHS(t)){s.maHocSinh=t;s.id=t;}
            else if(isDate(t)){s.ngaySinh=t;}
            else if(t==='Nam'){s.gioiTinh='Nam';}
            else if(t==='NÃ¡Â»Â¯'){s.gioiTinh='NÃ¡Â»Â¯';}
            else if(t.indexOf('@')!==-1){}
            else if(isIdOrCCCD(t)){}
            else if(t.indexOf('xÃƒÂ©t duyÃ¡Â»â€¡t')!==-1||t.indexOf('tiÃ¡ÂºÂ¿p nhÃ¡ÂºÂ­n')!==-1||t.indexOf('ChÃ¡Â»Â')!==-1){s.trangThai=t;}
            else if(isLop(t)){s.lop=t;afterLop=true;}
            else if(afterLop&&isNVSchool(t)){nvs.push(t.substring(0,80));}
        });
        for(var i=0;i<cells.length;i++){
            var t=cells[i];
            if(t&&t.length>2&&!t.startsWith('HS')&&!isDate(t)&&!isIdOrCCCD(t)&&
               t!=='Nam'&&t!=='NÃ¡Â»Â¯'&&t.indexOf('@')===-1&&t.indexOf('TrÃ†Â°Ã¡Â»Âng')===-1&&
               t.indexOf('THPT')===-1&&t.indexOf('PTDT')===-1&&!isLop(t)&&!allDigits(t)&&
               t.indexOf('duyÃ¡Â»â€¡t')===-1&&t.indexOf('nhÃ¡ÂºÂ­n')===-1&&t.indexOf('Thao')===-1){
                if(!s.hoTen) s.hoTen=t;
            }
        }
        if(nvs[0]) s.nv1=nvs[0];
        if(nvs[1]) s.nv2=nvs[1];
        if(nvs[2]) s.nv3=nvs[2];
        if(s.maHocSinh||s.hoTen) students.push(s);
    });
    return {students:students,rowCount:rows.length};
}
"""

def _fetch_tsdc_data_sync():
    """Chay Playwright trong thread rieng, tra ve dict thong ke."""
    result = {}
    def _in_thread():
        try:
            import asyncio
            from playwright.async_api import async_playwright
            async def _run():
                print('[TSDC] Khoi dong Playwright...', flush=True)
                async with async_playwright() as pw:
                    br   = await pw.chromium.launch(headless=True, slow_mo=0)
                    page = await br.new_page(viewport={'width':1400,'height':900})
                    # Login
                    print('[TSDC] Login...', flush=True)
                    await page.goto('https://qlts.tsdc.edu.vn', wait_until='domcontentloaded', timeout=20000)
                    await asyncio.sleep(1.5)
                    if await page.is_visible('input[type="password"]'):
                        await page.fill('input[type="text"]',  'qni_thcs_chuvanan1')
                        await page.fill('input[type="password"]', 'QuangNgai@06')
                        await page.click('button[type="submit"]')
                        await asyncio.sleep(4)
                    print('[TSDC] Navigate...', flush=True)
                    await page.goto('https://qlts.tsdc.edu.vn/quan-ly-ho-so', wait_until='networkidle', timeout=25000)
                    await asyncio.sleep(1.5)
                    await page.keyboard.press('Escape')
                    await asyncio.sleep(0.3)
                    # Click menu
                    print('[TSDC] Click menu...', flush=True)
                    for item in await page.query_selector_all('li, a'):
                        try:
                            txt = await item.inner_text()
                            if ('d\u1ef1 tuy\u1ec3n' in txt or 'du tuyen' in txt.lower()) and len(txt) < 50:
                                await item.click(); await asyncio.sleep(2.5); break
                        except Exception: pass
                    # No Escape here - it would close the filter
                    await asyncio.sleep(0.5)
                    # Select Cap 3
                    print('[TSDC] Chon Cap 3...', flush=True)
                    sels = await page.query_selector_all('.el-select')
                    for sel in sels:
                        try:
                            await sel.click(); await asyncio.sleep(0.5)
                            opts = page.locator('.el-select-dropdown:not([style*="display: none"]) .el-select-dropdown__item')
                            cnt  = await opts.count(); found = False
                            for j in range(cnt):
                                txt = await opts.nth(j).inner_text()
                                if 'c\u1ea5p 3' in txt.lower():
                                    await opts.nth(j).click(); found = True; break
                            if not found: await page.keyboard.press('Escape'); await asyncio.sleep(0.2)
                            else: break
                        except Exception: await page.keyboard.press('Escape')
                    await asyncio.sleep(1.2)  # wait for Dot dropdown to reload
                    # Select dot
                    print('[TSDC] Chon Dot THU...', flush=True)
                    sels = await page.query_selector_all('.el-select')
                    for sel in sels:
                        try:
                            await sel.click(); await asyncio.sleep(0.5)
                            opts = page.locator('.el-select-dropdown:not([style*="display: none"]) .el-select-dropdown__item')
                            cnt  = await opts.count(); found = False
                            for j in range(cnt):
                                txt = await opts.nth(j).inner_text()
                                if 'th\u1EED' in txt.lower() or 'th\u1EEC' in txt:  # (THá»¬) hoáº·c (thá»­)
                                    await opts.nth(j).click(); found = True; break
                            if not found: await page.keyboard.press('Escape'); await asyncio.sleep(0.2)
                            else: break
                        except Exception: await page.keyboard.press('Escape')
                    await asyncio.sleep(0.4)
                    # Search
                    print('[TSDC] Tim kiem...', flush=True)
                    btn = page.locator('button').filter(has_text='T\u00ecm ki\u1ebfm')
                    if await btn.count() > 0:
                        await btn.first.click(); await asyncio.sleep(5.5)
                    # Debug screenshot
                    await page.screenshot(path='C:/Users/HPZBook/Desktop/HS LOP 10/tsdc_debug_flask.png', full_page=False)
                    print('[TSDC] Screenshot saved!', flush=True)
                    # Extract
                    print('[TSDC] Extract data...', flush=True)
                    data = await page.evaluate(_TSDC_JS)
                    await br.close()
                    print(f'[TSDC] OK: {len(data.get("students",[]))} hoc sinh', flush=True)
                    return data.get('students', [])
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            students = loop.run_until_complete(_run())
            loop.close()
            # Aggregate stats
            status_map, nv1_map, nv2_map, nv3_map = {}, {}, {}, {}
            for s in students:
                st = s.get('trangThai', 'Kh\u00f4ng r\u00f5') or 'Kh\u00f4ng r\u00f5'
                status_map[st] = status_map.get(st, 0) + 1
                for field, mp in [('nv1', nv1_map), ('nv2', nv2_map), ('nv3', nv3_map)]:
                    v = s.get(field, '')
                    if v: mp[v] = mp.get(v, 0) + 1
            result['ok'] = {
                'total': len(students),
                'status': dict(sorted(status_map.items(), key=lambda x: -x[1])),
                'nv1': sorted(nv1_map.items(), key=lambda x: -x[1]),
                'nv2': sorted(nv2_map.items(), key=lambda x: -x[1]),
                'nv3': sorted(nv3_map.items(), key=lambda x: -x[1]),
                'students': [{'hoTen':s.get('hoTen',''),'lop':s.get('lop',''),
                              'trangThai':s.get('trangThai',''),
                              'nv1':s.get('nv1',''),'nv2':s.get('nv2',''),'nv3':s.get('nv3','')}
                             for s in students],
                'updated_at': datetime.now().strftime('%d/%m/%Y %H:%M:%S')
            }
        except Exception as e:
            import traceback; traceback.print_exc()
            result['err'] = str(e)
    t = threading.Thread(target=_in_thread, daemon=True)
    t.start(); t.join(timeout=150)
    return result.get('ok'), result.get('err', 'Timeout sau 150s - kiem tra ket noi mang')

@app.route('/api/tsdc-stats')
def api_tsdc_stats():
    import socket, json as _json
    is_pa = 'pythonanywhere' in socket.gethostname().lower() or os.path.exists('/home/ts102627')
    force = request.args.get('force', '0') == '1'
    now   = _time_mod.time()
    if is_pa:
        conn = get_db()
        row = conn.execute('SELECT data_json, pushed_at FROM tsdc_cache WHERE id=1').fetchone()
        conn.close()
        if row and row['data_json']:
            data = _json.loads(row['data_json'])
            data['pushed_at'] = row['pushed_at']; data['source'] = 'db'
            return jsonify({'success': True, **data})
        return jsonify({'error': 'Chua co du lieu. May local chua push len.'}), 404
    if not force and _tsdc_cache['data'] and now - _tsdc_cache['ts'] < _TSDC_TTL:
        return jsonify({'success': True, 'cached': True, **_tsdc_cache['data']})
    if _tsdc_cache['fetching']:
        if _tsdc_cache['data']:
            return jsonify({'success': True, 'cached': True, 'stale': True, **_tsdc_cache['data']})
        return jsonify({'loading': True, 'message': 'Dang ket noi TSDC...'}), 202
    def _bg_fetch():
        _tsdc_cache['fetching'] = True
        data, err = _fetch_tsdc_data_sync()
        if data: _tsdc_cache['data'] = data; _tsdc_cache['ts'] = _time_mod.time(); _tsdc_cache['err'] = None
        else: _tsdc_cache['err'] = err
        _tsdc_cache['fetching'] = False
    threading.Thread(target=_bg_fetch, daemon=True).start()
    return jsonify({'loading': True, 'message': 'Dang khoi dong...'}), 202

@app.route('/api/tsdc-status')
def api_tsdc_status():
    if _tsdc_cache.get('data'): return jsonify({'ready': True, **_tsdc_cache['data']})
    if _tsdc_cache.get('fetching'): return jsonify({'ready': False, 'loading': True})
    err = _tsdc_cache.get('err')
    if err: return jsonify({'ready': False, 'error': f'Loi: {err}'})
    return jsonify({'ready': False})

_TSDC_PUSH_TOKEN = os.environ.get('TSDC_PUSH_TOKEN', 'chuvanan_tsdc_push_2026')

@app.route('/api/tsdc-push', methods=['POST'])
def api_tsdc_push():
    try:
        import json as _json
        body = request.get_json(force=True, silent=True) or {}
        if body.get('token','') != _TSDC_PUSH_TOKEN:
            return jsonify({'error': 'Token khong hop le'}), 403
        data = {k: v for k, v in body.items() if k != 'token'}
        data_json = _json.dumps(data, ensure_ascii=False)
        pushed_at = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        conn = get_db()
        conn.execute('INSERT OR REPLACE INTO tsdc_cache (id,data_json,pushed_at,pushed_by) VALUES (1,?,?,?)',
                     (data_json, pushed_at, request.remote_addr))
        conn.commit(); conn.close()
        _tsdc_cache['data'] = data; _tsdc_cache['ts'] = _time_mod.time()
        total = data.get('total', 0)
        print(f'[TSDC-PUSH] {total} HS luc {pushed_at} tu {request.remote_addr}', flush=True)
        # Sau khi luu cache, tu dong sync vao bang students
        try:
            _tsdc_sync_students(data.get('students', []))
        except Exception as se:
            print(f'[TSDC-SYNC] Loi sync: {se}', flush=True)
        return jsonify({'success': True, 'total': total, 'pushed_at': pushed_at})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500

def _normalize_name(name):
    """Chuan hoa ten: bo dau, thuong, xoa khoang trang thua."""
    import unicodedata
    name = name.strip().lower()
    name = unicodedata.normalize('NFD', name)
    name = ''.join(c for c in name if unicodedata.category(c) != 'Mn')
    return ' '.join(name.split())

def _tsdc_sync_students(tsdc_students):
    """Khop hoc sinh TSDC voi DB theo:
      1. ma_hoso == tsdc_cccd.lstrip('0')  (chinh xac nhat - Excel ma_hoso = CCCD khong co so 0 dau)
      2. cccd (neu da luu truoc do)
      3. DOB + ten (fallback)
    """
    if not tsdc_students:
        return 0
    conn = get_db()
    updated = 0
    from database import migrate_db
    migrate_db()
    now_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')

    # Load all local students
    all_students = conn.execute(
        "SELECT id, ma_hoso, ho_ten, ho_ten_khong_dau, ngay_sinh, cccd FROM students"
    ).fetchall()

    # Build lookup maps
    by_ma_hoso   = {}   # ma_hoso stripped leading 0 → id
    by_cccd_exact = {}  # cccd stored in DB → id
    by_dob_name  = {}   # dob|name_norm → id (fallback)

    for s in all_students:
        # 1. Ma hoso (= CCCD cua HS, khong co so 0 dau)
        mhs = (s['ma_hoso'] or '').strip().lstrip('0')
        if mhs and mhs.isdigit():          # Chi map neu ma_hoso la so (khong phai dang 9B1_001)
            by_ma_hoso[mhs] = s['id']

        # 2. CCCD da luu (co the co so 0 dau hoac khong)
        if s['cccd']:
            by_cccd_exact[s['cccd'].strip().lstrip('0')] = s['id']

        # 3. DOB + ten (fallback)
        dob = (s['ngay_sinh'] or '').strip()
        name_with_space = _normalize_name(s['ho_ten'] or '')
        by_dob_name[dob + '|' + name_with_space] = s['id']
        if s['ho_ten_khong_dau']:
            by_dob_name[dob + '|' + _normalize_name(s['ho_ten_khong_dau'])] = s['id']

    for ts in tsdc_students:
        tsdc_cccd     = (ts.get('cccd') or '').strip()
        tsdc_cccd_raw = tsdc_cccd.lstrip('0')   # TSDC CCCD bo so 0 dau
        tsdc_ma_dd    = (ts.get('maDinhDanh') or '').strip()
        dob    = (ts.get('ngaySinh') or '').strip()
        name   = (ts.get('hoTen') or '').strip()
        nv1    = ts.get('nv1', '')
        nv2    = ts.get('nv2', '')
        nv3    = ts.get('nv3', '')
        trang  = ts.get('trangThai', '')
        mahoso = ts.get('maHocSinh', '')  # Ma ho so TSDC (HSO2651...)

        student_id = None
        match_by   = None

        # 1. Ma_hoso (chinh xac nhat): ma_hoso trong DB == CCCD TSDC bo so 0 dau
        if tsdc_cccd_raw and tsdc_cccd_raw in by_ma_hoso:
            student_id = by_ma_hoso[tsdc_cccd_raw]
            match_by   = f'ma_hoso={tsdc_cccd_raw}'

        # 2. CCCD da luu trong DB
        if not student_id and tsdc_cccd_raw and tsdc_cccd_raw in by_cccd_exact:
            student_id = by_cccd_exact[tsdc_cccd_raw]
            match_by   = f'cccd_stored={tsdc_cccd}'

        # 3. DOB + ten (fallback)
        if not student_id and dob:
            key = dob + '|' + _normalize_name(name)
            if key in by_dob_name:
                student_id = by_dob_name[key]
                match_by   = 'dob+name'

        if student_id:
            conn.execute("""UPDATE students
                SET tsdc_nv1=?, tsdc_nv2=?, tsdc_nv3=?,
                    tsdc_trang_thai=?, tsdc_ma_hoso=?, tsdc_updated_at=?,
                    cccd=COALESCE(NULLIF(?,  ''), cccd),
                    ma_dinh_danh_gd=COALESCE(NULLIF(?, ''), ma_dinh_danh_gd)
                WHERE id=?""",
                (nv1, nv2, nv3, trang, mahoso, now_str,
                 tsdc_cccd, tsdc_ma_dd, student_id))
            updated += 1
            print(f'[TSDC-SYNC] [{match_by}] {name} ({dob}) NV1={nv1[:30] if nv1 else "N/A"}', flush=True)
        else:
            print(f'[TSDC-SYNC] ✗ Khong khop: "{name}" ({dob}) CCCD={tsdc_cccd}', flush=True)

    conn.commit()
    conn.close()
    print(f'[TSDC-SYNC] Xong: {updated}/{len(tsdc_students)} hoc sinh cap nhat.', flush=True)
    return updated

@app.route('/api/tsdc-sync-students', methods=['POST'])
def api_tsdc_sync_students():
    """Goi thu cong de sync du lieu TSDC vao bang students."""
    body = request.get_json(force=True, silent=True) or {}
    if body.get('token', '') != _TSDC_PUSH_TOKEN:
        return jsonify({'error': 'Token khong hop le'}), 403
    students = _tsdc_cache.get('data', {}).get('students', []) if _tsdc_cache else []
    if not students:
        return jsonify({'error': 'Chua co du lieu TSDC trong cache'}), 404
    n = _tsdc_sync_students(students)
    return jsonify({'success': True, 'updated': n})

@app.route('/api/tsdc-debug', methods=['POST'])
def api_tsdc_debug():
    """Debug: hien thi trang thai DB + thu match TSDC students vao DB.
    Body: { token, students: [...TSDC data...] }
    """
    import unicodedata as _ud
    body = request.get_json(force=True, silent=True) or {}
    if body.get('token', '') != _TSDC_PUSH_TOKEN:
        return jsonify({'error': 'Token khong hop le'}), 403

    conn = get_db()
    rows = conn.execute(
        'SELECT id, ma_hoso, ho_ten, ho_ten_khong_dau, ngay_sinh, cccd, ma_dinh_danh_gd FROM students LIMIT 20'
    ).fetchall()
    conn.close()

    # Hien thi DB state
    db_sample = [dict(r) for r in rows]

    # Thu match
    tsdc_list = body.get('students', [])
    match_log = []
    if tsdc_list:
        def _norm(name):
            name = name.strip().lower()
            name = _ud.normalize('NFD', name)
            name = ''.join(c for c in name if _ud.category(c) != 'Mn')
            return ' '.join(name.split())

        conn2 = get_db()
        all_stu = conn2.execute(
            'SELECT id, ma_hoso, ho_ten, ho_ten_khong_dau, ngay_sinh, cccd FROM students'
        ).fetchall()
        conn2.close()

        by_ma_hoso, by_cccd, by_dob_name = {}, {}, {}
        for s in all_stu:
            mhs = (s['ma_hoso'] or '').strip().lstrip('0')
            if mhs and mhs.isdigit():
                by_ma_hoso[mhs] = {'id': s['id'], 'ho_ten': s['ho_ten']}
            if s['cccd']:
                by_cccd[s['cccd'].strip().lstrip('0')] = {'id': s['id'], 'ho_ten': s['ho_ten']}
            dob = (s['ngay_sinh'] or '').strip()
            n1 = _norm(s['ho_ten'] or '')
            by_dob_name[dob + '|' + n1] = {'id': s['id'], 'ho_ten': s['ho_ten']}
            if s['ho_ten_khong_dau']:
                by_dob_name[dob + '|' + _norm(s['ho_ten_khong_dau'])] = {'id': s['id'], 'ho_ten': s['ho_ten']}

        for ts in tsdc_list:
            cccd = (ts.get('cccd') or '').strip()
            cccd_raw = cccd.lstrip('0')
            dob = (ts.get('ngaySinh') or '').strip()
            name = (ts.get('hoTen') or '').strip()
            name_norm = _norm(name)
            key = dob + '|' + name_norm

            entry = {'tsdc_name': name, 'dob': dob, 'cccd': cccd,
                     'name_norm': name_norm, 'key': key,
                     'match': None, 'match_by': None}

            m = by_ma_hoso.get(cccd_raw)
            if m: entry['match'] = m; entry['match_by'] = f'ma_hoso({cccd_raw})'
            if not m:
                m = by_cccd.get(cccd_raw)
                if m: entry['match'] = m; entry['match_by'] = f'cccd_stored'
            if not m:
                m = by_dob_name.get(key)
                if m: entry['match'] = m; entry['match_by'] = f'dob+name'

            # Tim ngay sinh tuong ung trong DB
            dob_hits = [{'key': k, 'ho_ten': v['ho_ten']}
                        for k, v in by_dob_name.items() if dob in k and dob]
            entry['dob_hits_in_db'] = dob_hits[:5]
            match_log.append(entry)

    return jsonify({
        'success': True,
        'db_count': len(db_sample),
        'db_sample': db_sample,
        'match_log': match_log
    })

def create_dirs():
    """TÃ¡ÂºÂ¡o thÃ†Â° mÃ¡Â»Â¥c cÃ¡ÂºÂ§n thiÃ¡ÂºÂ¿t khi khÃ¡Â»Å¸i Ã„â€˜Ã¡Â»â„¢ng"""
    _base = os.environ.get('DATA_DIR', os.path.dirname(__file__))
    os.makedirs(os.path.join(_base, 'uploads'), exist_ok=True)
    os.makedirs(os.path.join(_base, 'backups'), exist_ok=True)

if __name__ == '__main__':
    create_dirs()
    init_db()
    app.run(host='0.0.0.0', port=5000, debug=False)
else:
    # ChÃ¡ÂºÂ¡y qua gunicorn (Render)
    create_dirs()
    init_db()




